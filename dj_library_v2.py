#!/usr/bin/env python3
"""
dj_library_v2.py — Biblioteca DJ
Discogs + Spotify: álbum-first matching, BPM via Deezer API,
gera XLSX e HTML interativo com views por LP e por Faixas.
"""

import sys, os, re, time, html as html_module, math, threading, webbrowser
import http.server, urllib.parse
from pathlib import Path

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")

# ==============================================================================
# 0. DEPENDÊNCIAS
# ==============================================================================
import subprocess

def pip_install(*pkgs):
    subprocess.check_call([sys.executable, "-m", "pip", "install", "-q", *pkgs])

try:
    import discogs_client
except ImportError:
    print("Instalando discogs-client..."); pip_install("discogs-client")
    import discogs_client

try:
    import spotipy
    from spotipy.oauth2 import SpotifyOAuth
except ImportError:
    print("Instalando spotipy..."); pip_install("spotipy")
    import spotipy
    from spotipy.oauth2 import SpotifyOAuth

try:
    from thefuzz import fuzz
except ImportError:
    print("Instalando thefuzz..."); pip_install("thefuzz", "python-Levenshtein")
    from thefuzz import fuzz

try:
    from unidecode import unidecode
except ImportError:
    print("Instalando unidecode..."); pip_install("unidecode")
    from unidecode import unidecode

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Instalando openpyxl..."); pip_install("openpyxl")
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter

import pandas as pd
import json, requests

# ==============================================================================
# 1. CONFIGURAÇÃO
# ==============================================================================
DISCOGS_USER   = "amsa2diop"
DISCOGS_TOKEN  = "FmUsfbDTXUFCniDqrTckrwBfJxvIljOGuMdygfVD"
SP_CLIENT_ID   = "1ab6d898c52d42a19b737f451ce31e2a"
SP_CLIENT_SEC  = "3c8b2f47049b44e2af6937ea835e1f2f"
SP_REDIRECT    = "http://127.0.0.1:1410/"
SP_SCOPE       = "playlist-modify-public playlist-modify-private"

LIMIAR_ACEITO   = 72
LIMIAR_REVISAR  = 55
LIMIAR_ALBUM    = 58   # score mínimo para aceitar match de álbum

VA_KEYWORDS = {"various", "v.a.", "varios", "variados", "aa.vv.", "various artists"}

WORK_DIR = Path(__file__).parent


# ==============================================================================
# 2. AUTENTICAÇÃO SPOTIFY
# ==============================================================================
def spotify_auth():
    sp_oauth = SpotifyOAuth(
        client_id     = SP_CLIENT_ID,
        client_secret = SP_CLIENT_SEC,
        redirect_uri  = SP_REDIRECT,
        scope         = SP_SCOPE,
        cache_path    = str(WORK_DIR / ".spotify_cache"),
        open_browser  = False,
    )
    token_info = sp_oauth.get_cached_token()
    if token_info and not sp_oauth.is_token_expired(token_info):
        print("✓ Token Spotify em cache (válido).")
        return spotipy.Spotify(auth_manager=sp_oauth)

    auth_code_holder = [None]

    class OAuthHandler(http.server.BaseHTTPRequestHandler):
        def do_GET(self):
            parsed = urllib.parse.urlparse(self.path)
            params = urllib.parse.parse_qs(parsed.query)
            if "code" in params:
                auth_code_holder[0] = params["code"][0]
                self.send_response(200)
                self.send_header("Content-type", "text/html; charset=utf-8")
                self.end_headers()
                self.wfile.write(b"<h2>Autenticado. Pode fechar esta aba.</h2>")
            else:
                self.send_response(400); self.end_headers()
        def log_message(self, *a): pass

    server = http.server.HTTPServer(("127.0.0.1", 1410), OAuthHandler)
    t = threading.Thread(target=server.handle_request, daemon=True)
    t.start()
    auth_url = sp_oauth.get_authorize_url()
    print(f"\nAbrindo navegador para autenticação Spotify...")
    webbrowser.open(auth_url)
    t.join(timeout=120)
    server.server_close()
    if not auth_code_holder[0]:
        raise TimeoutError("Autenticação não completada em 120s.")
    sp_oauth.get_access_token(auth_code_holder[0])
    return spotipy.Spotify(auth_manager=sp_oauth)


# ==============================================================================
# 3. UTILITÁRIOS
# ==============================================================================
def normalize(text):
    if not text: return ""
    text = unidecode(str(text)).lower()
    text = re.sub(r"\(.*?\)|\[.*?\]", "", text)
    text = re.sub(r"\s*[-–—]\s+.*$", "", text)
    text = re.sub(r"\bfeat\.?\b.*$|\bft\.?\b.*$", "", text, flags=re.IGNORECASE)
    text = re.sub(r"[^a-z0-9 ]", "", text)
    return text.strip()


def match_score(disc_track, disc_artist, sp_track, sp_artist, album_artist=""):
    dt = normalize(disc_track)
    da = normalize(disc_artist)
    aa = normalize(album_artist)
    st = normalize(sp_track)
    sa = normalize(sp_artist)

    score_track  = fuzz.token_sort_ratio(dt, st)
    score_artist = max(
        fuzz.token_sort_ratio(da, sa),
        fuzz.token_sort_ratio(aa, sa) if aa else 0,
    )
    score = score_track * 0.65 + score_artist * 0.35

    sp_lo = sp_track.lower()
    di_lo = disc_track.lower()
    if re.search(r"\blive\b|ao vivo|concert|en vivo", sp_lo) and \
       not re.search(r"live|ao vivo|concert", di_lo):
        score -= 20
    if re.search(r"\bremix\b|\bedit\b|rework|dub mix", sp_lo) and \
       not re.search(r"remix|edit|rework|dub", di_lo):
        score -= 15
    if re.search(r"remaster", sp_lo) and not re.search(r"remaster", di_lo):
        score -= 5
    return max(0, score)


def is_va(artist_name):
    return normalize(artist_name) in VA_KEYWORDS or "various" in normalize(artist_name)


def key_to_camelot(key, mode):
    major = ["8B","3B","10B","5B","12B","7B","2B","9B","4B","11B","6B","1B"]
    minor = ["5A","12A","7A","2A","9A","4A","11A","6A","1A","8A","3A","10A"]
    if key is None or (isinstance(key, float) and math.isnan(key)):
        return ""
    return major[int(key)] if mode == 1 else minor[int(key)]


def card_colors(raw_hex, fator=0.28):
    """Retorna (bg_hex, css_vars) para gradiente do card com texto legível.
    Usa luminância da cor RAW (não clareada) porque o gradiente usa a cor raw."""
    if not raw_hex or len(raw_hex) != 7:
        return "", ""
    try:
        r, g, b = int(raw_hex[1:3], 16), int(raw_hex[3:5], 16), int(raw_hex[5:7], 16)
        rp = int(r + (255 - r) * fator)
        gp = int(g + (255 - g) * fator)
        bp = int(b + (255 - b) * fator)
        bg = f"#{rp:02X}{gp:02X}{bp:02X}"
        # Luminância da cor raw (o gradiente usa a cor raw, não a clareada)
        lum = (0.299 * r + 0.587 * g + 0.114 * b) / 255
        if lum < 0.55:  # cor raw escura/média → gradiente escuro → texto branco
            cvars = ("--text:#F2F2F2;--text2:#E5E5E5;--text3:#CCCCCC;"
                     "--bdr:rgba(255,255,255,.18);--bdr2:rgba(255,255,255,.12);"
                     "--bpm-col:#F2F2F2;--tracks-bg:rgba(0,0,0,.28);"
                     "--acc2:rgba(255,255,255,.4);"
                     "--cef-inp-bg:#252525;--cef-inp-tc:#e6e6e6")
        else:  # cor raw clara → gradiente claro → texto preto
            cvars = ("--text:#111;--text2:#333;--text3:#555;"
                     "--bdr:rgba(0,0,0,.13);--bdr2:rgba(0,0,0,.08);"
                     "--tracks-bg:rgba(255,255,255,.65);"
                     "--cef-inp-bg:#e9e5de;--cef-inp-tc:#1a1a1a")
        return bg, cvars
    except Exception:
        return "", ""


def safe_float(v):
    try:
        f = float(v)
        return None if math.isnan(f) else f
    except: return None


def normalize_bpm(bpm):
    """Divide BPM > 140 by 2 repeatedly to land in 70-140 range."""
    if bpm is None: return None
    b = float(bpm)
    while b > 140:
        b /= 2
    return round(b, 1) if b > 0 else None


def esc(v):
    if v is None or (isinstance(v, float) and math.isnan(v)): return ""
    return html_module.escape(str(v))

_EMPTY = {"nan", "None", "none", ""}

def _clean(val):
    """Normalise any falsy / nan-string value to empty string."""
    s = str(val) if val is not None else ""
    return "" if s in _EMPTY else s

def _decade(year_int):
    """Return decade chip key for a year integer."""
    if not year_int: return ""
    if year_int < 1970: return "pre70"
    if year_int < 1980: return "70s"
    if year_int < 1990: return "80s"
    if year_int < 2000: return "90s"
    if year_int < 2010: return "2000s"
    if year_int < 2020: return "2010s"
    return "2020s"

def _genre_style(genres_s, styles_s, limit=5):
    """Return deduplicated 'Genre · Style' string."""
    parts = [p.strip() for p in (str(genres_s or "") + "," + str(styles_s or "")).split(",")
             if p.strip() and p.strip() != "nan"]
    seen: set = set()
    return " · ".join(p for p in parts if not (p in seen or seen.add(p)))[:limit * 30]


# ==============================================================================
# 4. DISCOGS: COLETA DE ÁLBUNS E FAIXAS
# ==============================================================================
def get_discogs_collection():
    print("Coletando coleção do Discogs...")
    d    = discogs_client.Client("DJLibrary/2.0", user_token=DISCOGS_TOKEN)
    user = d.user(DISCOGS_USER)
    col  = user.collection_folders[0].releases

    albums, tracks = [], []
    total, done = col.count, 0

    for item in col:
        try:
            rel = item.release
            album_artist = rel.artists[0].name if rel.artists else "Unknown"
            album_artist = re.sub(r"\s*\(\d+\)$", "", album_artist)
            album_title  = rel.title
            release_id   = rel.id
            year         = getattr(rel, "year", None)
            genres       = ", ".join(getattr(rel, "genres", []) or [])
            styles       = ", ".join(getattr(rel, "styles", []) or [])
            images       = getattr(rel, "images", []) or []
            cover_url    = images[0].get("uri", "") if images else ""
            thumb_url    = images[0].get("uri150", cover_url) if images else ""

            albums.append(dict(
                release_id=release_id, album_artist=album_artist,
                album_title=album_title, year=year, genres=genres,
                styles=styles, cover_url=cover_url, thumb_url=thumb_url,
            ))

            for trk in rel.tracklist:
                if trk.type_ != "track" or not trk.position:
                    continue
                trk_artist = trk.artists[0].name if trk.artists else album_artist
                trk_artist = re.sub(r"\s*\(\d+\)$", "", trk_artist)
                tracks.append(dict(
                    release_id=release_id, album_artist=album_artist,
                    album_title=album_title, year=year, genres=genres,
                    styles=styles, cover_url=cover_url, thumb_url=thumb_url,
                    position=trk.position, track_title=trk.title,
                    artist_raw=trk.artists[0].name if trk.artists else album_artist,
                    artist_clean=trk_artist,
                ))

            done += 1
            print(f"\r  {done}/{total} albums | {len(tracks)} tracks", end="", flush=True)
            time.sleep(0.5)

        except Exception as e:
            done += 1
            print(f"\n  Aviso album #{done}: {e}")
            time.sleep(2)

    print(f"\n✓ {done} albums | {len(tracks)} tracks")
    return pd.DataFrame(albums), pd.DataFrame(tracks)


def fetch_collection_ids_rest():
    """
    Retorna set de release_ids (int) da coleção via REST API.
    Rápido: pagina o índice da coleção sem buscar detalhes de cada release.
    """
    session = requests.Session()
    session.headers.update({
        "Authorization": f"Discogs token={DISCOGS_TOKEN}",
        "User-Agent":    "DJLibrary/2.0",
    })
    ids = set()
    page = 1
    while True:
        try:
            r = session.get(
                f"https://api.discogs.com/users/{DISCOGS_USER}/collection/folders/0/releases",
                params={"per_page": 100, "page": page, "sort": "added", "sort_order": "desc"},
                timeout=15,
            )
            if r.status_code == 429:
                print("  Rate limit — aguardando 60s..."); time.sleep(60); continue
            if r.status_code != 200:
                print(f"  Erro HTTP {r.status_code} ao listar coleção"); break
            data = r.json()
            for item in data.get("releases", []):
                ids.add(item["basic_information"]["id"])
            pages = data.get("pagination", {}).get("pages", 1)
            if page >= pages:
                break
            page += 1
            time.sleep(0.5)
        except Exception as e:
            print(f"  Erro ao buscar coleção: {e}"); break
    return ids


def fetch_releases_details(release_ids):
    """
    Busca detalhes de faixas para uma lista de release_ids via REST API.
    Retorna (albums_df, tracks_df) no mesmo formato de get_discogs_collection().
    """
    session = requests.Session()
    session.headers.update({
        "Authorization": f"Discogs token={DISCOGS_TOKEN}",
        "User-Agent":    "DJLibrary/2.0",
    })
    albums, tracks = [], []
    total = len(release_ids)

    for i, rid in enumerate(release_ids):
        try:
            r = session.get(f"https://api.discogs.com/releases/{rid}", timeout=15)
            if r.status_code == 429:
                print(f"\n  Rate limit — aguardando 60s..."); time.sleep(60)
                r = session.get(f"https://api.discogs.com/releases/{rid}", timeout=15)
            if r.status_code != 200:
                print(f"\n  Aviso: release {rid} → HTTP {r.status_code}")
                time.sleep(1); continue

            data = r.json()
            raw_artist   = (data.get("artists") or [{}])[0].get("name", "Unknown")
            album_artist = re.sub(r"\s*\(\d+\)$", "", raw_artist)
            album_title  = data.get("title", "")
            year         = data.get("year") or None
            genres       = ", ".join(data.get("genres") or [])
            styles       = ", ".join(data.get("styles") or [])
            images       = data.get("images") or []
            cover_url    = images[0].get("uri", "") if images else ""
            thumb_url    = images[0].get("uri150", cover_url) if images else ""

            albums.append(dict(
                release_id=rid, album_artist=album_artist, album_title=album_title,
                year=year, genres=genres, styles=styles,
                cover_url=cover_url, thumb_url=thumb_url,
            ))

            for trk in data.get("tracklist") or []:
                if trk.get("type_", "track") not in ("track", "") or not trk.get("position"):
                    continue
                raw_trk_artist = (trk.get("artists") or [{}])[0].get("name", album_artist)
                trk_artist = re.sub(r"\s*\(\d+\)$", "", raw_trk_artist)
                tracks.append(dict(
                    release_id=rid, album_artist=album_artist, album_title=album_title,
                    year=year, genres=genres, styles=styles,
                    cover_url=cover_url, thumb_url=thumb_url,
                    position=trk["position"], track_title=trk.get("title", ""),
                    artist_raw=raw_trk_artist, artist_clean=trk_artist,
                ))

        except Exception as e:
            print(f"\n  Aviso release {rid}: {e}")

        print(f"\r  {i+1}/{total} releases | {len(tracks)} faixas", end="", flush=True)
        time.sleep(0.6)

    print()
    return pd.DataFrame(albums) if albums else pd.DataFrame(), \
           pd.DataFrame(tracks) if tracks else pd.DataFrame()


# ==============================================================================
# 5. SPOTIFY: ALBUM-FIRST MATCHING
# ==============================================================================
def sp_search(sp, q, stype="track", limit=5):
    try:
        res = sp.search(q=q, type=stype, limit=limit)
        items = res[stype + "s"]["items"]
        return items or []
    except Exception:
        time.sleep(1)
        return []


def find_spotify_album(sp, album_title, album_artist, year=None):
    """Busca o álbum no Spotify. Retorna (album_obj, score) ou (None, 0)."""
    # Constrói variantes do nome do artista (para casos como "Igor Babidi" → "Babidi")
    words = album_artist.split()
    artist_variants = list(dict.fromkeys([
        album_artist,
        words[-1] if len(words) > 1 else None,       # última palavra
        words[0] if len(words) > 1 else None,         # primeira palavra
        " ".join(words[1:]) if len(words) > 2 else None,  # sem primeira palavra
    ]))
    artist_variants = [v for v in artist_variants if v]

    queries = []
    for av in artist_variants:
        queries.append(f'album:"{album_title}" artist:"{av}"')
    queries.append(f'{album_artist} {album_title}')
    if year:
        queries.insert(1, f'album:"{album_title}" artist:"{album_artist}" year:{year}')

    best, best_score = None, 0

    for q in queries:
        time.sleep(0.25)
        albums = sp_search(sp, q, stype="album", limit=5)
        for alb in albums:
            sp_title  = alb.get("name", "")
            sp_artist = alb["artists"][0]["name"] if alb.get("artists") else ""
            sc_title  = fuzz.token_sort_ratio(normalize(album_title), normalize(sp_title))
            sc_artist = max(
                fuzz.token_sort_ratio(normalize(av), normalize(sp_artist))
                for av in artist_variants
            )
            score = sc_title * 0.6 + sc_artist * 0.4
            if score > best_score:
                best_score = score
                best = alb
        if best_score >= 75:
            break

    return (best, best_score) if best_score >= LIMIAR_ALBUM else (None, 0)


def get_album_tracks(sp, album_id):
    """Retorna lista completa de faixas de um álbum Spotify."""
    tracks = []
    try:
        page = sp.album_tracks(album_id, limit=50)
        while page:
            tracks.extend(page["items"])
            page = sp.next(page) if page.get("next") else None
    except Exception:
        pass
    return tracks


def match_disc_to_album(disc_tracks_df, sp_tracks, album_artist):
    """
    Tenta fazer match das faixas Discogs com as faixas do álbum Spotify.
    Retorna dict {idx: match_dict} e lista de índices não encontrados.
    """
    matched, unmatched = {}, []

    for idx, row in disc_tracks_df.iterrows():
        best_score, best_sp = 0, None
        for sp_t in sp_tracks:
            sp_name   = sp_t.get("name", "")
            sp_artist = sp_t["artists"][0]["name"] if sp_t.get("artists") else ""
            # Dentro do álbum: peso maior para título
            sc_title  = fuzz.token_sort_ratio(normalize(row["track_title"]), normalize(sp_name))
            sc_artist = fuzz.token_sort_ratio(normalize(row["artist_clean"]), normalize(sp_artist))
            sc = sc_title * 0.80 + sc_artist * 0.20
            # Penalizações
            if re.search(r"\blive\b|ao vivo|concert", sp_name.lower()) and \
               not re.search(r"live|ao vivo|concert", row["track_title"].lower()):
                sc -= 20
            if re.search(r"\bremix\b|\bedit\b|rework", sp_name.lower()) and \
               not re.search(r"remix|edit|rework", row["track_title"].lower()):
                sc -= 15
            sc = max(0, sc)
            if sc > best_score:
                best_score, best_sp = sc, sp_t

        if best_sp and best_score >= LIMIAR_ACEITO:
            matched[idx] = dict(
                spotify_uri     = best_sp["uri"],
                found_name      = best_sp["name"],
                found_artist    = best_sp["artists"][0]["name"] if best_sp.get("artists") else "",
                track_id        = best_sp["id"],
                match_score     = round(best_score, 1),
                search_strategy = "album_first",
            )
        else:
            unmatched.append(idx)

    return matched, unmatched


def search_track_fallback(sp, artist, album_artist, title):
    """Busca individual de faixa — fallback para faixas não encontradas no álbum."""
    time.sleep(0.3)

    def run(q, limit=5):
        return sp_search(sp, q, stype="track", limit=limit)

    items = run(f'artist:"{artist}" track:"{title}"')
    if not items: time.sleep(0.2); items = run(f"{artist} {title}")
    if not items and artist.lower() != album_artist.lower():
        time.sleep(0.2); items = run(f"{album_artist} {title}")
    if not items: time.sleep(0.2); items = run(f'track:"{title}"')

    empty = dict(spotify_uri=None, found_name=None, found_artist=None,
                 track_id=None, match_score=0.0, search_strategy="not_found")
    if not items: return empty

    best_score, best = -1, empty
    for item in items[:5]:
        sp_name   = item.get("name", "")
        sp_artist = item["artists"][0]["name"] if item.get("artists") else ""
        sc = match_score(title, artist, sp_name, sp_artist, album_artist)
        if sc > best_score:
            best_score = sc
            best = dict(
                spotify_uri     = item["uri"],
                found_name      = sp_name,
                found_artist    = sp_artist,
                track_id        = item["id"],
                match_score     = round(sc, 1),
                search_strategy = "track",
            )
    return best


def run_album_first_matching(sp, df):
    """
    Matching completo com lógica álbum-first.
    - Para álbuns não-V.A.: primeiro busca o álbum Spotify, depois faz
      match das faixas nele. Faixas não encontradas no álbum fazem
      busca individual como fallback.
    - Para V.A./coletâneas: busca individual para cada faixa.
    """
    sp_cols = ["spotify_uri","found_name","found_artist","track_id",
               "match_score","search_strategy"]
    for col in sp_cols:
        df[col] = None

    albums = df.groupby("release_id")
    total_albums = len(albums)
    done_albums  = 0
    done_tracks  = 0

    for release_id, group in albums:
        first      = group.iloc[0]
        alb_artist = str(first["album_artist"])
        alb_title  = str(first["album_title"])
        year       = first.get("year")
        n_tracks   = len(group)

        if is_va(alb_artist):
            # V.A.: busca individual
            for idx, row in group.iterrows():
                res = search_track_fallback(sp, row["artist_clean"], alb_artist, row["track_title"])
                for k, v in res.items():
                    df.at[idx, k] = v
                done_tracks += 1
        else:
            # Álbum normal: tenta album-first
            sp_album, alb_score = find_spotify_album(sp, alb_title, alb_artist, year)

            if sp_album:
                sp_tracks = get_album_tracks(sp, sp_album["id"])
                matched, unmatched = match_disc_to_album(group, sp_tracks, alb_artist)

                for idx, res in matched.items():
                    for k, v in res.items():
                        df.at[idx, k] = v
                    done_tracks += 1

                # Fallback individual para não encontrados no álbum
                for idx in unmatched:
                    row = df.loc[idx]
                    res = search_track_fallback(sp, row["artist_clean"], alb_artist, row["track_title"])
                    for k, v in res.items():
                        df.at[idx, k] = v
                    done_tracks += 1
            else:
                # Álbum não encontrado: busca individual para todas as faixas
                for idx, row in group.iterrows():
                    res = search_track_fallback(sp, row["artist_clean"], alb_artist, row["track_title"])
                    for k, v in res.items():
                        df.at[idx, k] = v
                    done_tracks += 1

        done_albums += 1
        print(f"\r  {done_albums}/{total_albums} albums | {done_tracks}/{len(df)} faixas",
              end="", flush=True)

    print(f"\n✓ Matching concluído")

    df["status"] = df["match_score"].apply(
        lambda s: "ACEITO"   if pd.notna(s) and s >= LIMIAR_ACEITO
        else ("REVISAR"      if pd.notna(s) and s >= LIMIAR_REVISAR
        else "REJEITADO")
    )
    return df


# ==============================================================================
# 6. BPM — Deezer API (gratuita) + Spotify fallback
# ==============================================================================
_DEEZER_SESSION = None

def _deezer_session():
    global _DEEZER_SESSION
    if _DEEZER_SESSION is None:
        _DEEZER_SESSION = requests.Session()
        _DEEZER_SESSION.headers.update({
            "User-Agent": "DJLibrary/2.0",
            "Accept": "application/json",
        })
    return _DEEZER_SESSION


def _deezer_search_bpm(artist, title):
    """
    Busca BPM no Deezer por artista+título.
    Retorna (bpm_float, deezer_id_str) ou (None, None).
    """
    q = f"{normalize(artist)} {normalize(title)}"
    try:
        r = _deezer_session().get(
            "https://api.deezer.com/search",
            params={"q": q, "limit": 5},
            timeout=12,
        )
        if r.status_code == 429:
            time.sleep(10)
            return None, None
        if r.status_code != 200:
            return None, None
        hits = r.json().get("data", [])
        if not hits:
            return None, None

        # Escolhe a melhor correspondência por fuzzy
        t_norm = normalize(title)
        a_norm = normalize(artist)
        best_id = None
        best_sc = 0
        for hit in hits:
            sc_t = fuzz.token_sort_ratio(t_norm, normalize(hit.get("title", "")))
            sc_a = fuzz.token_sort_ratio(a_norm, normalize(hit.get("artist", {}).get("name", "")))
            sc = sc_t * 0.7 + sc_a * 0.3
            if sc > best_sc:
                best_sc = sc
                best_id = hit["id"]

        if best_sc < 60 or not best_id:
            return None, None

        # Busca detalhes da track (tem BPM)
        r2 = _deezer_session().get(f"https://api.deezer.com/track/{best_id}", timeout=12)
        if r2.status_code != 200:
            return None, None
        bpm = r2.json().get("bpm", 0)
        if bpm and float(bpm) > 0:
            return normalize_bpm(float(bpm)), str(best_id)
        return None, None
    except Exception:
        return None, None


def get_bpm_deezer(df):
    """
    Busca BPM via Deezer API (gratuita, sem autenticação) para faixas aceitas.
    Usa found_artist + found_name (nomes do Spotify) como query de busca.
    Retorna dict {track_id: {bpm, source}}.
    """
    accepted = df[df["status"] == "ACEITO"].copy()
    # Agrupa por track_id para não buscar duplicatas
    unique_tracks = accepted.dropna(subset=["track_id"]).drop_duplicates(subset=["track_id"])
    total = len(unique_tracks)
    if total == 0:
        return {}

    print(f"\nBuscando BPM no Deezer para {total} faixas únicas...")
    results = {}
    found = 0

    for i, (_, row) in enumerate(unique_tracks.iterrows()):
        tid = str(row.get("track_id", "") or "")
        artist = str(row.get("found_artist", "") or row.get("album_artist", "") or "")
        title  = str(row.get("found_name", "") or row.get("track_title", "") or "")

        bpm, deezer_id = _deezer_search_bpm(artist, title)
        if bpm:
            results[tid] = {"bpm": bpm, "deezer_id": deezer_id or "", "source": "deezer"}
            found += 1

        if (i + 1) % 50 == 0 or (i + 1) == total:
            print(f"\r  {i+1}/{total} | com BPM: {found}", end="", flush=True)
        time.sleep(0.25)  # ~4 req/s, bem abaixo do limite do Deezer

    print(f"\r  {total}/{total} | com BPM: {found} ({100*found//max(total,1)}%)")
    if found == 0:
        print("  ⚠ Deezer não retornou BPM — verifique conexão.")
    return results


def get_bpm_spotify(sp, track_ids):
    """Tenta buscar BPM via Spotify audio_features (pode falhar com 403)."""
    valid   = [t for t in track_ids if t]
    results = {}
    chunks  = [valid[i:i+100] for i in range(0, len(valid), 100)]

    for chunk in chunks:
        time.sleep(0.4)
        try:
            feats = sp.audio_features(chunk)
            for f in feats or []:
                if f:
                    results[f["id"]] = {
                        "bpm":          round(f["tempo"], 1),
                        "energy":       round(f["energy"], 2),
                        "danceability": round(f["danceability"], 2),
                        "valence":      round(f["valence"], 2),
                        "key":          f["key"],
                        "mode":         f["mode"],
                        "camelot":      key_to_camelot(f["key"], f["mode"]),
                        "source":       "spotify",
                    }
        except Exception as e:
            if "403" in str(e):
                print("\n  ✗ Spotify audio_features: 403 — app sem acesso (depreciado para novos apps).")
                return results  # Não tenta mais
            pass
    return results


def fetch_bpm(sp, df):
    """
    Busca BPM apenas para faixas aceitas que ainda NÃO têm BPM.
    Estratégia: 1) Spotify audio_features  2) Deezer API (gratuita)
    """
    bpm_cols = ["bpm","energy","danceability","valence","key","mode","camelot","deezer_id"]
    for col in bpm_cols:
        if col not in df.columns:
            df[col] = None

    # Apenas faixas aceitas SEM BPM
    _bpm_safe = df["bpm"].apply(safe_float)
    missing_mask = (df["status"] == "ACEITO") & _bpm_safe.isna()
    missing_ids  = df[missing_mask]["track_id"].dropna().unique().tolist()

    if not missing_ids:
        print(f"✓ BPM completo — {_bpm_safe.notna().sum()} faixas, nenhuma nova para buscar.")
        return df

    print(f"\nBuscando BPM para {len(missing_ids)} faixas sem BPM...")

    # 1. Tenta Spotify audio_features
    print(f"  Tentando Spotify audio_features...")
    bpm_map = get_bpm_spotify(sp, missing_ids)

    # 2. Para o que o Spotify não cobriu, usa Deezer
    missing_after_sp = [tid for tid in missing_ids if tid not in bpm_map]
    if missing_after_sp:
        # Passa apenas as linhas sem BPM para get_bpm_deezer
        df_missing = df[df["track_id"].astype(str).isin(missing_after_sp)].copy()
        deezer_map = get_bpm_deezer(df_missing)
        bpm_map.update(deezer_map)

    # Aplica BPM ao dataframe (apenas linhas afetadas)
    for i, row in df[missing_mask].iterrows():
        tid = str(row.get("track_id") or "")
        if tid and tid in bpm_map:
            entry = bpm_map[tid]
            df.at[i, "bpm"] = entry.get("bpm")
            for k in ["energy","danceability","valence","key","mode","camelot","deezer_id"]:
                if k in entry:
                    df.at[i, k] = entry.get(k)

    newly_found = df.loc[missing_mask, "bpm"].apply(safe_float).notna().sum()
    total_found = df["bpm"].apply(safe_float).notna().sum()
    print(f"✓ BPM novo: {newly_found}/{len(missing_ids)} | total: {total_found}")
    return df


# ==============================================================================
# 7. EXECUÇÃO PRINCIPAL
# ==============================================================================
def main():
    sp = spotify_auth()
    print("✓ Spotify autenticado.\n")

    backup_tracks_path = WORK_DIR / "backup_tracks.csv"
    backup_v2_path     = WORK_DIR / "backup_matched_v2.csv"
    backup_bpm_path    = WORK_DIR / "backup_bpm.csv"
    BPM_COLS = ["bpm","energy","danceability","valence","key","mode","camelot","deezer_id"]

    # ── 1. Coleção Discogs (incremental) ─────────────────────────────────────
    print("Verificando coleção no Discogs...")
    current_ids = fetch_collection_ids_rest()      # set de release_id (int) — rápido
    print(f"  {len(current_ids)} releases na coleção atual")

    if backup_tracks_path.exists():
        df_existing  = pd.read_csv(backup_tracks_path)
        existing_ids = set(df_existing["release_id"].dropna().astype(int).unique())
        new_ids      = current_ids - existing_ids
        removed_ids  = existing_ids - current_ids

        if not new_ids and not removed_ids:
            print(f"  ✓ Sem alterações ({len(existing_ids)} releases, {len(df_existing)} faixas)")
            df = df_existing
        else:
            if removed_ids:
                print(f"  {len(removed_ids)} release(s) removido(s) da coleção")
                df_existing = df_existing[~df_existing["release_id"].isin(removed_ids)]
            if new_ids:
                print(f"  {len(new_ids)} novo(s) release(s) — buscando faixas via REST...")
                _, df_new = fetch_releases_details(list(new_ids))
                df = pd.concat([df_existing, df_new], ignore_index=True) if not df_new.empty else df_existing
            else:
                df = df_existing
            df.to_csv(backup_tracks_path, index=False)
            print(f"  ✓ backup_tracks.csv atualizado")
    else:
        print("  Primeira execução — buscando coleção completa...")
        _, df = get_discogs_collection()
        df.to_csv(backup_tracks_path, index=False)

    print(f"  {df['release_id'].nunique()} albums | {len(df)} faixas\n")

    # ── 2. Spotify matching (incremental) ─────────────────────────────────────
    if backup_v2_path.exists():
        df_matched   = pd.read_csv(backup_v2_path)
        matched_ids  = set(df_matched["release_id"].dropna().astype(int).unique())
        all_ids      = set(df["release_id"].dropna().astype(int).unique())
        new_to_match = all_ids - matched_ids

        if new_to_match:
            print(f"Matching {len(new_to_match)} novo(s) release(s) no Spotify...")
            df_new_tracks  = df[df["release_id"].isin(new_to_match)].copy()
            df_new_matched = run_album_first_matching(sp, df_new_tracks)
            df_matched = pd.concat([df_matched, df_new_matched], ignore_index=True)
            df_matched.to_csv(backup_v2_path, index=False)
            print(f"  ✓ backup_matched_v2.csv atualizado")
        else:
            print(f"✓ Spotify matching completo ({len(df_matched)} faixas, sem novos releases)")
    elif (WORK_DIR / "backup_final.csv").exists():
        # Compatibilidade com versão anterior
        print("Carregando backup legado (backup_final.csv)...")
        df_matched = pd.read_csv(WORK_DIR / "backup_final.csv")
        df_matched.to_csv(backup_v2_path, index=False)   # migra para v2
    else:
        print("Executando Spotify matching completo...")
        df_matched = run_album_first_matching(sp, df.copy())
        df_matched.to_csv(backup_v2_path, index=False)
        print(f"  ✓ Salvo: {backup_v2_path.name}")

    aceitos   = (df_matched["status"] == "ACEITO").sum()
    revisao   = (df_matched["status"] == "REVISAR").sum()
    rejeitado = (df_matched["status"] == "REJEITADO").sum()
    print(f"  Aceitos: {aceitos} | Revisar: {revisao} | Rejeitados: {rejeitado}\n")

    # ── 3. BPM (incremental) ──────────────────────────────────────────────────
    for col in BPM_COLS:
        if col not in df_matched.columns:
            df_matched[col] = None

    # Carrega backup BPM e preenche o DataFrame
    if backup_bpm_path.exists():
        bpm_df  = pd.read_csv(backup_bpm_path).dropna(subset=["track_id"])
        bpm_df["bpm"] = bpm_df["bpm"].apply(lambda v: normalize_bpm(safe_float(v)))
        bpm_map = {str(r["track_id"]): r for _, r in bpm_df.iterrows()}
        for i, row in df_matched.iterrows():
            tid = str(row.get("track_id") or "")
            if tid and tid in bpm_map:
                for col in BPM_COLS:
                    df_matched.at[i, col] = bpm_map[tid].get(col)
        filled = df_matched["bpm"].apply(safe_float).notna().sum()
        print(f"✓ BPM carregado: {filled} faixas do backup")

    # Busca BPM apenas para faixas aceitas sem BPM (novas ou nunca processadas)
    df_matched = fetch_bpm(sp, df_matched)

    # Salva backup BPM (combina existente + novo)
    bpm_found = df_matched[df_matched["bpm"].apply(safe_float).notna()]
    if len(bpm_found) > 0:
        bpm_found[["track_id"] + BPM_COLS].to_csv(backup_bpm_path, index=False)

    return sp, df_matched


# ==============================================================================
# 8. XLSX
# ==============================================================================
def generate_xlsx(df):
    print("\nGerando XLSX...")
    path = WORK_DIR / "MinhaColecao_DJ.xlsx"
    wb   = openpyxl.Workbook()

    hdr_fill  = PatternFill("solid", fgColor="7B3020")
    hdr_font  = Font(color="FAF7F3", bold=True)
    ok_fill   = PatternFill("solid", fgColor="DDF0E8")
    warn_fill = PatternFill("solid", fgColor="FDF3DC")
    rej_fill  = PatternFill("solid", fgColor="F5E0DC")

    cols = ["status","album_artist","artist_clean","track_title","album_title",
            "year","genres","styles","bpm","camelot","energy","danceability",
            "valence","match_score","found_artist","found_name",
            "search_strategy","spotify_uri","cover_url"]

    def write_sheet(ws, data, name):
        ws.title = name
        if data.empty:
            ws.append(["Nenhum dado."]); return
        sub = data[[c for c in cols if c in data.columns]].copy()
        ws.append(list(sub.columns))
        for cell in ws[1]:
            cell.fill = hdr_fill; cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center")
        status_idx = list(sub.columns).index("status") + 1 if "status" in sub.columns else None
        for _, row in sub.iterrows():
            ws.append([None if isinstance(v, float) and math.isnan(v) else v for v in row])
        if status_idx:
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                val  = row[status_idx - 1].value
                fill = ok_fill if val=="ACEITO" else (warn_fill if val=="REVISAR" else rej_fill)
                for cell in row: cell.fill = fill
        for i, col in enumerate(ws.columns, 1):
            max_len = max((len(str(c.value or "")) for c in col), default=8)
            ws.column_dimensions[get_column_letter(i)].width = min(max_len + 2, 50)

    ws1 = wb.active
    write_sheet(ws1, df.sort_values(["album_artist","album_title","position"],
                na_position="last"), "Todos (por album)")

    ws2 = wb.create_sheet()
    df_ok = df[df["status"]=="ACEITO"].copy()
    df_ok["bpm"] = pd.to_numeric(df_ok["bpm"], errors="coerce")
    write_sheet(ws2, df_ok.sort_values("bpm", na_position="last"), "Aceitos (por BPM)")

    ws3 = wb.create_sheet()
    write_sheet(ws3, df[df["status"]=="REVISAR"].sort_values("match_score", ascending=False),
                "Para Revisar")

    ws4 = wb.create_sheet()
    write_sheet(ws4, df[df["status"]=="REJEITADO"], "Nao Encontrados")

    wb.save(path)
    print(f"✓ {path.name} salvo")
    return path


# ==============================================================================
# 9. PLAYLIST SPOTIFY
# ==============================================================================
def create_playlist(sp, df):
    print("\nCriando playlist Spotify (por BPM)...")
    df_pl = df[df["status"]=="ACEITO"].copy()
    df_pl["bpm"] = pd.to_numeric(df_pl["bpm"], errors="coerce")
    df_pl = df_pl.sort_values("bpm", na_position="last")
    uris  = df_pl["spotify_uri"].dropna().tolist()
    if not uris:
        print("  Nenhuma faixa aceita para playlist."); return

    from datetime import datetime
    uid = sp.current_user()["id"]
    pl  = sp.user_playlist_create(
        user=uid, name="Meu Discogs — por BPM", public=False,
        description=f"Vinil ordenado por BPM | {datetime.now().strftime('%d/%m/%Y')}",
    )
    pid = pl["id"]
    chunks = [uris[i:i+100] for i in range(0, len(uris), 100)]
    for i, chunk in enumerate(chunks):
        sp.playlist_add_items(pid, chunk); time.sleep(0.3)
        print(f"\r  Lote {i+1}/{len(chunks)}", end="", flush=True)
    print(f"\n✓ Playlist criada com {len(uris)} faixas")
    pl_url = f"https://open.spotify.com/playlist/{pid}"
    (WORK_DIR / "backup_playlist.txt").write_text(pl_url, encoding="utf-8")
    print(f"  {pl_url}")


# ==============================================================================
# 10. HTML — DESIGN BEGE/TERROSO, DUAS VIEWS
# ==============================================================================
CSS = """
:root{
  --bg:#FFFFFF;--bg2:#F5F1EB;--card:#FAF8F5;--card2:#F0EBE3;
  --text:#111;--text2:#444;--text3:#777;
  --acc:#111;--acc2:#444;
  --bdr:#DDD;--bdr2:#E8E8E8;
  --bpm-col:#111;
  --r:14px;--r-sm:8px;
  --shadow:0 2px 12px rgba(0,0,0,.06),0 1px 3px rgba(0,0,0,.03);
  --shadow-h:0 4px 20px rgba(0,0,0,.10),0 2px 6px rgba(0,0,0,.05);
}
*{margin:0;padding:0;box-sizing:border-box}
body{font-family:system-ui,-apple-system,"Segoe UI",Arial,sans-serif;background:var(--bg);color:var(--text);min-height:100vh}
h1,h2,h3,.serif{font-family:Georgia,"Times New Roman",serif}

/* HEADER — dashboard single-line */
.site-header{background:var(--bg);border-bottom:1px solid var(--bdr);
  display:flex;align-items:center;gap:1.4rem;padding:0 2.5rem;
  height:54px;position:sticky;top:0;z-index:100;flex-shrink:0}
.logo{display:flex;align-items:center;gap:.45rem;text-decoration:none;flex-shrink:0}
.logo-name{font-family:Georgia,serif;font-weight:normal;color:var(--acc);
  font-size:1.4rem;letter-spacing:.04em;white-space:nowrap}
.header-sep{width:1px;height:22px;background:var(--bdr);flex-shrink:0}
.site-stats{display:flex;gap:.55rem;align-items:center;flex:1;min-width:0}
.stat-item{font-size:.74rem;color:var(--text2);white-space:nowrap;font-weight:500}
.stat-item strong{color:var(--text);font-weight:700}
.stat-dot{color:var(--bdr2);font-size:.62rem}
.header-tabs{display:flex;margin-left:auto;flex-shrink:0}
.tab-btn{padding:0 1.1rem;height:54px;font-size:.68rem;letter-spacing:.12em;
  text-transform:uppercase;cursor:pointer;background:none;border:none;
  border-bottom:2px solid transparent;color:var(--text3);
  transition:color .18s,border-color .18s;white-space:nowrap}
.tab-btn.active{color:var(--acc);border-bottom-color:var(--acc)}
.tab-btn:hover:not(.active){color:var(--text2)}

/* COPY BADGE */
.copy-badge{display:inline-block;padding:.1rem .42rem;border-radius:4px;font-size:.6rem;
  font-weight:700;letter-spacing:.07em;text-transform:uppercase;
  background:#fff;color:#111;border:1px solid rgba(0,0,0,.22);
  margin-left:.35rem;vertical-align:middle}


/* CUSTOM FIELDS */
.fields-row{display:flex;flex-wrap:wrap;gap:.25rem .7rem;margin-top:.45rem}
.field-item{font-size:.65rem;color:var(--text3)}
.field-item strong{color:var(--text2);font-weight:600}
.field-notes{font-size:.68rem;color:var(--text2);font-style:italic;
  margin-top:.3rem;padding:.28rem .5rem;background:var(--tracks-bg);
  border-left:2px solid var(--bdr);border-radius:0 4px 4px 0}

/* VIEW */
.view{display:none}.view.active{display:block}

/* CONTROLS — sticky below header */
.controls{position:sticky;top:54px;z-index:50;
  background:rgba(255,255,255,.96);backdrop-filter:blur(10px);
  border-bottom:1px solid var(--bdr);padding:.5rem 2.5rem;
  display:flex;flex-direction:column;gap:.38rem}
/* back-to-top button inside controls row */
.back-top-btn{margin-left:auto;width:30px;height:30px;border-radius:8px;
  border:1px solid var(--bdr);background:var(--bg2);
  display:flex;align-items:center;justify-content:center;
  cursor:pointer;color:var(--text3);font-size:.85rem;flex-shrink:0;
  transition:background .15s,color .15s}
.back-top-btn:hover{background:var(--acc);color:#fff;border-color:var(--acc)}
/* incomplete link in results-bar */
.incomplete-link{font-size:.72rem;color:#bbb;text-decoration:none;
  border-bottom:1px dashed #ccc;cursor:pointer;transition:color .15s,border-color .15s}
.incomplete-link:hover{color:#555;border-color:#888}
.h-flag{font-size:.62rem;font-weight:700;border-radius:4px;padding:2px 6px;
  flex-shrink:0;text-transform:uppercase;letter-spacing:.04em;white-space:nowrap}
.f-bpm {background:#FFF3E0;color:#BF5700}
.f-sp  {background:#E3F2FD;color:#1258A8}
.f-both{background:#FCE4EC;color:#880E4F}
/* Inc badges — visible only when #grid-faixas has inc-mode class */
.inc-badge{display:none}
#grid-faixas.inc-mode .inc-badge{display:inline-block;margin-top:.2rem}
.incomplete-link.active{color:#555;border-color:#888;font-weight:600}
.ctrl-row{display:flex;gap:.45rem;align-items:center;flex-wrap:wrap}
.ctrl-input{background:var(--bg2);border:1px solid var(--bdr);color:var(--text);
  padding:.4rem .9rem;border-radius:20px;font-size:.84rem;outline:none;
  flex:1;min-width:160px;transition:border-color .15s}
.ctrl-input:focus{border-color:var(--acc)}
.ctrl-sel{background:var(--bg2);border:1px solid var(--bdr);color:var(--text);
  padding:.4rem .7rem;border-radius:8px;font-size:.79rem;outline:none;cursor:pointer}
.ctrl-btn{background:transparent;border:1px solid var(--bdr);color:var(--text2);
  padding:.36rem .8rem;border-radius:8px;font-size:.72rem;letter-spacing:.04em;
  text-transform:uppercase;cursor:pointer;white-space:nowrap;transition:all .18s}
.ctrl-btn:hover{border-color:var(--acc);color:var(--acc)}

/* CHIPS */
.chip{padding:.26rem .62rem;border-radius:20px;font-size:.69rem;letter-spacing:.03em;
  cursor:pointer;border:1px solid var(--bdr);background:transparent;color:var(--text3);
  transition:all .15s;white-space:nowrap}
.chip.active{background:#F5F1EB;border-color:#999;color:#111;font-weight:600}
.chip.multi-active{background:#EAE5DA;border-color:#888;color:#111;font-weight:600}
.chip:hover:not(.active):not(.multi-active){border-color:var(--acc2);color:var(--acc)}
.chips-label{font-size:.62rem;color:var(--text3);white-space:nowrap;letter-spacing:.05em;
  text-transform:uppercase}

/* PLAYLIST BUTTON */
.playlist-btn{display:inline-flex;align-items:center;gap:.35rem;padding:.32rem .8rem;
  border-radius:20px;font-size:.7rem;font-weight:600;letter-spacing:.04em;
  background:#1DB954;color:#fff;border:none;cursor:pointer;text-decoration:none;
  transition:opacity .15s;white-space:nowrap}
.playlist-btn:hover{opacity:.85}

/* SOCIAL BUTTONS (header) — links sem contorno */
.header-social-btn{display:inline-flex;align-items:center;gap:.32rem;font-size:.74rem;
  color:var(--text2);text-decoration:none;padding:.1rem .2rem;font-weight:500;
  transition:color .15s;white-space:nowrap;flex-shrink:0}
.header-social-btn:hover{color:var(--acc)}

/* LOGO LINK */
.logo-link{text-decoration:none;color:inherit;transition:opacity .15s}
.logo-link:hover{opacity:.7}

/* FILTER PANEL */
.filter-toggle-btn{display:inline-flex;align-items:center;gap:.3rem;
  border:1px solid var(--bdr);background:transparent;color:var(--text3);
  padding:.32rem .8rem;border-radius:8px;font-size:.72rem;letter-spacing:.04em;
  text-transform:uppercase;cursor:pointer;white-space:nowrap;transition:all .18s;flex-shrink:0}
.filter-toggle-btn.open{border-color:var(--acc2);color:var(--acc)}
.filter-panel{display:none;flex-direction:column;gap:.32rem;padding:.18rem 0}
.filter-panel.open{display:flex}
.filter-group{display:flex;align-items:center;gap:.45rem;flex-wrap:wrap}
.filter-group-label{font-size:.6rem;color:var(--text3);text-transform:uppercase;
  letter-spacing:.07em;white-space:nowrap;flex-basis:100%;flex-shrink:0;margin-bottom:.05rem}

/* MAIN */
.main{max-width:1200px;margin:0 auto;padding:1.4rem 2.5rem}
.results-bar{font-size:.7rem;color:var(--text3);letter-spacing:.06em;
  text-transform:uppercase;margin-bottom:1rem}

/* ── LP VIEW ── */
.albums-grid{display:flex;flex-direction:column;gap:.45rem}
.album-card{background:var(--card);border:1px solid var(--bdr);border-radius:var(--r);
  overflow:hidden;transition:box-shadow .2s;box-shadow:var(--shadow);position:relative}
.album-card:hover{box-shadow:var(--shadow-h)}
.album-card.hidden{display:none}
.cover-blur{position:absolute;inset:0;background-size:cover;background-position:center;
  filter:blur(40px) saturate(3) brightness(.5);opacity:.09;transform:scale(1.15);pointer-events:none}
.album-header{display:flex;align-items:center;gap:1rem;padding:1rem 1.2rem;
  cursor:pointer;user-select:none;position:relative;z-index:1}
.cover-wrap{flex-shrink:0}
.cover-img{width:90px;height:90px;object-fit:cover;border-radius:var(--r-sm);display:block;
  box-shadow:0 2px 8px rgba(0,0,0,.18)}
.cover-ph{width:90px;height:90px;border-radius:var(--r-sm);background:var(--bg2);
  display:flex;align-items:center;justify-content:center;color:var(--text3);font-size:1.8rem}
.album-info{flex:1;min-width:0}
.alb-artist{font-weight:600;font-size:.94rem;color:var(--text)}
.alb-title{color:var(--text2);font-size:.81rem;margin-top:.1rem;
  white-space:nowrap;overflow:hidden;text-overflow:ellipsis}
.alb-meta{font-size:.67rem;color:var(--text3);margin-top:.3rem;
  line-height:1.45;overflow:hidden;display:-webkit-box;-webkit-line-clamp:2;-webkit-box-orient:vertical}
.alb-tracks-info{font-size:.65rem;color:var(--text3);margin-top:.22rem}
.alb-date-added{font-size:.58rem;color:var(--text3,#aaa);margin-top:.1rem;opacity:.7}
.toggle-btn{background:none;border:none;color:var(--text3);cursor:pointer;
  padding:.3rem .6rem;font-size:.85rem;flex-shrink:0;transition:transform .25s,color .2s;
  position:relative;z-index:1}
.album-card.open .toggle-btn{transform:rotate(180deg);color:var(--acc)}
.tracks-list{border-top:1px solid var(--bdr2);background:var(--tracks-bg,rgba(255,255,255,.75))}
.tracks-list.collapsed{display:none}

/* TRACK ITEM (LP view) */
.track{padding:.58rem 1.2rem;border-bottom:1px solid var(--bdr2);
  display:flex;align-items:center;gap:.65rem;transition:background .1s}
.track:last-child{border-bottom:none}
.track:hover{background:rgba(0,0,0,.03)}
.trk-pos{font-size:.61rem;color:var(--text3);width:22px;flex-shrink:0;letter-spacing:.03em}
.trk-info{flex:1;min-width:100px}
.trk-artist{font-size:.67rem;color:var(--text3)}
.trk-name{font-size:.87rem;font-weight:500;color:var(--text)}
.trk-badges{display:flex;gap:.3rem;align-items:center;flex-shrink:0}
.badge-bpm{padding:.15rem .42rem;border-radius:5px;font-size:.71rem;font-weight:700;
  background:rgba(255,255,255,.93);color:#111;border:1px solid rgba(0,0,0,.16)}
.trk-btn-col{display:flex;flex-direction:column;gap:.2rem;align-items:stretch;flex-shrink:0;min-width:66px}
.trk-btn{font-size:.61rem;font-weight:600;color:#1A1A1A;background:rgba(255,255,255,.93);
  border:1px solid rgba(0,0,0,.16);border-radius:5px;padding:.18rem .4rem;
  text-align:center;text-decoration:none;white-space:nowrap;
  transition:all .15s;cursor:pointer;display:block}
.trk-btn:hover{border-color:#555;background:#eee}
.trk-status{width:5px;height:5px;border-radius:50%;flex-shrink:0;margin-left:.15rem}
.s-ok{background:#3A9060}.s-rev{background:#C08020}.s-rej{background:var(--bdr)}

/* EMBED */
.sp-embed{border-radius:var(--r-sm);display:block}
/* LP accordion: embed abaixo da linha de faixa */
.tracks-list .embed-below{padding:.45rem 1.2rem .6rem;border-top:1px solid var(--bdr2)}

/* ── TRACK VIEW (compact) ── */
.track-rows{display:flex;flex-direction:column;gap:0}
.track-row{display:flex;align-items:center;gap:8px;padding:5px 10px;
  border-bottom:1px solid rgba(0,0,0,.055);flex-wrap:wrap;
  position:relative;cursor:pointer;transition:filter .1s}
.track-row:last-child{border-bottom:none}
.track-row:hover{filter:brightness(.96)}
.track-row.hidden{display:none}
/* BPM — left of cover */
.c-bpm{font-size:.88rem;font-weight:700;min-width:38px;text-align:right;
  flex-shrink:0;letter-spacing:-.5px;color:var(--tc,#111);line-height:1}
.c-bpm small{display:block;font-size:.55rem;font-weight:500;letter-spacing:.04em;
  opacity:.55;text-transform:uppercase;margin-top:1px}
.c-bpm-none{min-width:38px;text-align:right;flex-shrink:0;
  color:var(--tc,#111);opacity:.2;font-size:.8rem}
/* Cover */
.c-ph{width:40px;height:40px;border-radius:5px;flex-shrink:0;
  display:flex;align-items:center;justify-content:center;
  font-size:1.1rem;color:rgba(255,255,255,.55)}
.c-thumb{width:40px;height:40px;border-radius:5px;object-fit:cover;flex-shrink:0}
/* Info */
.c-info{flex:1;min-width:0}
.c-title{font-size:.82rem;font-weight:600;white-space:nowrap;
  overflow:hidden;text-overflow:ellipsis;line-height:1.35;color:var(--tc,#111)}
.c-artist{font-size:.72rem;white-space:nowrap;overflow:hidden;
  text-overflow:ellipsis;line-height:1.3;color:var(--tc,#111);opacity:.72}
/* Inline extra info (badges + notes) */
.c-extra{display:flex;flex-wrap:wrap;align-items:center;gap:.18rem .3rem;margin-top:.18rem;min-width:0}
.c-badge{font-size:.58rem;font-weight:700;border-radius:3px;padding:1px 5px;
  letter-spacing:.04em;white-space:nowrap;
  background:rgba(255,255,255,.78);border:1px solid rgba(0,0,0,.1)}
.c-badge-trocar{color:#1258A8}
.c-badge-disc{color:#2E7D32}
.c-badge-nrec{color:#BF5700}
.c-notes-inline{font-size:.63rem;font-style:italic;color:var(--tc,#888);
  opacity:.75;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;min-width:0}
/* Icon-only buttons */
.c-btns{display:flex;gap:3px;flex-shrink:0}
.ic-btn{width:30px;height:30px;border-radius:8px;
  border:1px solid rgba(0,0,0,.10);background:rgba(255,255,255,.82);
  display:flex;align-items:center;justify-content:center;
  cursor:pointer;text-decoration:none;color:#444;flex-shrink:0;
  transition:background .1s;padding:0}
.ic-btn:hover{background:rgba(255,255,255,.97)}
.ic-btn svg{width:15px;height:15px;display:block}
.ic-btn[disabled]{opacity:.22;cursor:default;pointer-events:none}
/* Expandable details panel */
.c-details{flex-basis:100%;display:none;
  padding:.4rem .25rem .2rem;border-top:1px solid rgba(0,0,0,.07);margin-top:.15rem}
.c-details.open{display:block}
.c-details .tr-album{font-size:.73rem;font-style:italic;
  color:var(--tc,#555);opacity:.8;margin-bottom:.2rem;
  white-space:nowrap;overflow:hidden;text-overflow:ellipsis}
.c-details .tr-meta{font-size:.68rem;color:var(--tc,#777);opacity:.7}
/* Embed player inside details inherits row gradient */
.c-details .embed-below{padding:.4rem 0 0;
  border-top:1px solid rgba(255,255,255,.15);margin-top:.3rem}

/* CARD BUTTONS (LP view) */
.btn-card{display:inline-flex;align-items:center;justify-content:center;gap:.3rem;
  font-size:.67rem;font-weight:600;color:#1A1A1A;background:rgba(255,255,255,.93);
  border:1px solid rgba(0,0,0,.16);border-radius:6px;padding:.26rem .68rem;
  text-decoration:none;white-space:nowrap;transition:border-color .15s,color .15s}
.btn-card:hover{color:var(--acc);border-color:var(--acc2)}
.rg-col-center{display:flex;flex-direction:column;align-items:stretch;gap:.3rem;
  flex-shrink:0;position:relative;z-index:1}

/* RESPONSIVE */
@media(max-width:640px){
  .site-header{padding:0 1rem;gap:.8rem}
  .site-stats{display:none}
  .header-sep{display:none}
  .header-social-btn{display:none}
  .controls{padding:.45rem .9rem;top:48px}.main{padding:.9rem .9rem}
  .album-header{gap:.7rem;padding:.8rem .9rem}
  .cover-img,.cover-ph{width:68px;height:68px}
  /* Compact track rows */
  .track-row{padding:4px 8px;gap:7px}
  .c-ph,.c-thumb{width:36px;height:36px}
  .c-bpm{min-width:32px;font-size:.82rem}
  .ic-btn{width:27px;height:27px}
  .ic-btn svg{width:13px;height:13px}
}

/* ── EDIT MODE ──────────────────────────────────────────────────────────────── */
/* Pencil toggle button — lives beside filter/back-top buttons in ctrl-row */
.pencil-mode-btn{width:30px;height:30px;border-radius:8px;
  border:1px solid var(--bdr);background:var(--bg2);
  display:flex;align-items:center;justify-content:center;
  cursor:pointer;font-size:.9rem;flex-shrink:0;
  transition:background .15s,color .15s,border-color .15s;color:var(--text3)}
.pencil-mode-btn:hover{background:var(--acc);color:#fff;border-color:var(--acc)}
.pencil-mode-btn.active{background:#e2e2e2;border-color:#999;color:#333;font-weight:600}
/* Edit button on card — styled same as Discogs/Spotify btn-card, hidden until edit mode */
body:not(.edit-mode) .cef-edit-btn{display:none}
/* Inline edit form */
.card-edit-form{display:none;padding:.6rem 1rem .8rem;border-top:1px solid var(--bdr2);
  background:var(--tracks-bg,rgba(245,242,237,.92));color:var(--text,#111)}
.card-edit-form.open{display:block}
.cef-grid{display:grid;grid-template-columns:auto 1fr;gap:.28rem .65rem;
  align-items:center;margin-bottom:.5rem}
.cef-label{font-size:.65rem;font-weight:600;color:var(--text,#111);opacity:.6;
  text-transform:uppercase;letter-spacing:.06em;white-space:nowrap}
.cef-input{background:var(--cef-inp-bg,#e9e5de);border:1px solid var(--bdr);
  color:var(--cef-inp-tc,#1a1a1a);
  padding:.3rem .55rem;border-radius:6px;font-size:.8rem;outline:none;
  transition:border-color .15s;font-family:inherit;width:100%;box-sizing:border-box}
.cef-input:focus{border-color:var(--acc,#7A5C2C)}
select.cef-input option{background:#1e1e1e;color:#e6e6e6}
.cef-textarea{resize:vertical;min-height:46px}
.cef-actions{display:flex;align-items:center;gap:.5rem;flex-wrap:wrap;margin-top:.3rem}
.cef-status{font-size:.72rem;color:var(--text,#111);opacity:.7;flex:1}
.cef-btn{padding:.28rem .7rem;border-radius:6px;font-size:.74rem;cursor:pointer;
  border:1px solid var(--bdr,rgba(0,0,0,.15));transition:all .15s}
.cef-save{background:#1a1a1a;color:#fff;border-color:#1a1a1a}
.cef-save:hover{background:#3a3a3a}
.cef-cancel{background:transparent;color:var(--text,#555);opacity:.75;border-color:transparent}
.cef-cancel:hover{opacity:1;border-color:var(--bdr)}
/* Setup modal */
.setup-overlay{position:fixed;inset:0;background:rgba(0,0,0,.45);
  z-index:9000;display:none;align-items:center;justify-content:center}
.setup-overlay.open{display:flex}
.setup-box{background:var(--bg);border:1px solid var(--bdr);border-radius:12px;
  padding:1.6rem;max-width:400px;width:90%;box-shadow:0 8px 30px rgba(0,0,0,.18)}
.setup-box h3{margin-bottom:1rem;font-size:.95rem}
.setup-box label{display:block;font-size:.68rem;font-weight:600;color:var(--text3);
  text-transform:uppercase;letter-spacing:.05em;margin-bottom:.2rem;margin-top:.6rem}
.setup-box label:first-of-type{margin-top:0}
.setup-box a{color:var(--acc);font-size:.7rem}
.setup-input{width:100%;background:var(--bg2);border:1px solid var(--bdr);
  color:var(--text);padding:.38rem .7rem;border-radius:7px;font-size:.84rem;
  outline:none;box-sizing:border-box;font-family:inherit}
.setup-input:focus{border-color:var(--acc)}
#setup-status{font-size:.75rem;min-height:1.2rem;color:var(--text3);
  margin:.55rem 0 .3rem}
.setup-actions{display:flex;gap:.5rem;margin-top:.7rem}
.setup-btn-connect{background:#111;color:#fff;border:none;padding:.4rem .9rem;
  border-radius:7px;font-size:.8rem;cursor:pointer;flex:1}
.setup-btn-connect:hover{background:#333}
.setup-btn-cancel{background:transparent;border:1px solid var(--bdr);color:var(--text3);
  padding:.4rem .9rem;border-radius:7px;font-size:.8rem;cursor:pointer}
"""

JS = r"""
// ── VIEW SWITCHING ────────────────────────────────────────────────────────────
function switchView(v){
  document.querySelectorAll('.view').forEach(function(el){el.classList.remove('active')});
  document.querySelectorAll('.tab-btn').forEach(function(el){el.classList.remove('active')});
  document.getElementById('view-'+v).classList.add('active');
  document.querySelector('[data-v="'+v+'"]').classList.add('active');
  syncAllChips();
}

// ── SHARED FILTER STATE ───────────────────────────────────────────────────────
var origemFilter=new Set();
var nacionalFilter='all';
var decadeFilter=new Set();
var compilFilter='all';
var djFilter='all';
var paFilter='all';
var dupFilter='all';
var recebidoFilter='all';
var incFilterActive=false;

function syncAllChips(){
  document.querySelectorAll('.nac-chip').forEach(function(c){
    c.classList.toggle('active',c.dataset.val===nacionalFilter);
  });
  document.querySelectorAll('.decade-chip').forEach(function(c){
    if(c.dataset.val==='all'){c.classList.toggle('active',decadeFilter.size===0);c.classList.remove('multi-active');}
    else{c.classList.toggle('multi-active',decadeFilter.has(c.dataset.val));c.classList.remove('active');}
  });
  document.querySelectorAll('.compil-chip').forEach(function(c){
    c.classList.toggle('active',c.dataset.val===compilFilter);
  });
  document.querySelectorAll('.bpm-chip').forEach(function(c){
    if(c.dataset.val==='all'){c.classList.toggle('active',activeBpmFilter.size===0);c.classList.remove('multi-active');}
    else{c.classList.toggle('multi-active',activeBpmFilter.has(c.dataset.val));c.classList.remove('active');}
  });
  document.querySelectorAll('.origem-chip').forEach(function(c){
    if(c.dataset.val==='all'){c.classList.toggle('active',origemFilter.size===0);c.classList.remove('multi-active');}
    else{c.classList.toggle('multi-active',origemFilter.has(c.dataset.val));c.classList.remove('active');}
  });
}

// ── FILTER PANEL ──────────────────────────────────────────────────────────────
function toggleFilterPanel(vid){
  var fp=document.getElementById('fp-'+vid);
  var btn=document.getElementById('fp-btn-'+vid);
  var isOpen=fp.classList.toggle('open');
  if(btn){btn.classList.toggle('open',isOpen);btn.textContent=isOpen?'⚙ Filtros ▴':'⚙ Filtros ▾';}
  if(isOpen)syncAllChips();
}

// ── ALBUM ACCORDION ───────────────────────────────────────────────────────────
function toggleAlbum(hdr){
  var c=hdr.closest('.album-card'),l=c.querySelector('.tracks-list');
  c.classList.toggle('open');l.classList.toggle('collapsed');
}
function expandAll(){
  document.querySelectorAll('#grid-lp .album-card:not(.hidden)').forEach(function(c){
    c.classList.add('open');c.querySelector('.tracks-list').classList.remove('collapsed');
  });
}
function collapseAll(){
  document.querySelectorAll('#grid-lp .album-card').forEach(function(c){
    c.classList.remove('open');c.querySelector('.tracks-list').classList.add('collapsed');
  });
}

// ── FILTER FUNCTIONS ──────────────────────────────────────────────────────────
function setOrigemFilter(val,el){
  if(val==='all'){
    origemFilter.clear();
    document.querySelectorAll('.origem-chip').forEach(function(c){c.classList.remove('active');c.classList.remove('multi-active')});
    document.querySelectorAll('.origem-chip[data-val="all"]').forEach(function(c){c.classList.add('active')});
  }else{
    document.querySelectorAll('.origem-chip[data-val="all"]').forEach(function(c){c.classList.remove('active')});
    if(origemFilter.has(val)){
      origemFilter.delete(val);
      document.querySelectorAll('.origem-chip[data-val="'+CSS.escape(val)+'"]').forEach(function(c){c.classList.remove('multi-active')});
    }else{
      origemFilter.add(val);
      document.querySelectorAll('.origem-chip[data-val="'+CSS.escape(val)+'"]').forEach(function(c){c.classList.add('multi-active')});
    }
    if(origemFilter.size===0){document.querySelectorAll('.origem-chip[data-val="all"]').forEach(function(c){c.classList.add('active')});}
  }
  filterLP();filterTracks();
}
function setNacionalFilter(val,el){
  nacionalFilter=val;
  document.querySelectorAll('.nac-chip').forEach(function(c){c.classList.remove('active')});
  document.querySelectorAll('.nac-chip[data-val="'+val+'"]').forEach(function(c){c.classList.add('active')});
  filterLP();filterTracks();
}
function setDecadeFilter(val,el){
  if(val==='all'){
    decadeFilter.clear();
    document.querySelectorAll('.decade-chip').forEach(function(c){c.classList.remove('active');c.classList.remove('multi-active')});
    document.querySelectorAll('.decade-chip[data-val="all"]').forEach(function(c){c.classList.add('active')});
  } else {
    document.querySelectorAll('.decade-chip[data-val="all"]').forEach(function(c){c.classList.remove('active')});
    if(decadeFilter.has(val)){
      decadeFilter.delete(val);
      document.querySelectorAll('.decade-chip[data-val="'+val+'"]').forEach(function(c){c.classList.remove('multi-active')});
    } else {
      decadeFilter.add(val);
      document.querySelectorAll('.decade-chip[data-val="'+val+'"]').forEach(function(c){c.classList.add('multi-active')});
    }
    if(decadeFilter.size===0){
      document.querySelectorAll('.decade-chip[data-val="all"]').forEach(function(c){c.classList.add('active')});
    }
  }
  filterLP();filterTracks();
}
function setCompilFilter(val,el){
  compilFilter=val;
  document.querySelectorAll('.compil-chip').forEach(function(c){c.classList.remove('active')});
  document.querySelectorAll('.compil-chip[data-val="'+val+'"]').forEach(function(c){c.classList.add('active')});
  filterLP();filterTracks();
}
function setDjFilter(val,el){
  djFilter=val;
  document.querySelectorAll('.dj-chip').forEach(function(c){c.classList.remove('active')});
  document.querySelectorAll('.dj-chip[data-val="'+val+'"]').forEach(function(c){c.classList.add('active')});
  filterLP();filterTracks();
}
function setPaFilter(val,el){
  paFilter=val;
  document.querySelectorAll('.pa-chip').forEach(function(c){c.classList.remove('active')});
  document.querySelectorAll('.pa-chip[data-val="'+val+'"]').forEach(function(c){c.classList.add('active')});
  filterLP();filterTracks();
}
function setDupFilter(val,el){
  dupFilter=val;
  document.querySelectorAll('.dup-chip').forEach(function(c){c.classList.remove('active')});
  el.classList.add('active');
  filterLP();
}
function setRecebidoFilter(val,el){
  recebidoFilter=val;
  document.querySelectorAll('.recebido-chip').forEach(function(c){c.classList.remove('active')});
  document.querySelectorAll('.recebido-chip[data-val="'+val+'"]').forEach(function(c){c.classList.add('active')});
  filterLP();filterTracks();
}

// ── LP FILTER ─────────────────────────────────────────────────────────────────
function albumBpmOk(card, filterSet){
  if(filterSet.size===0)return true;
  var list=(card.dataset.bpmList||'').split(',').filter(Boolean).map(Number);
  return Array.from(filterSet).some(function(f){
    if(f==='with')return list.length>0;
    if(f==='nobpm')return list.length===0;
    return list.some(function(bpm){
      if(f==='sub70')return bpm<70;
      if(f==='140plus')return bpm>=140;
      var p=f.split('-');return bpm>=+p[0]&&bpm<+p[1];
    });
  });
}
function filterLP(){
  var q=document.getElementById('q-lp').value.toLowerCase().trim();
  var s=document.getElementById('sort-lp').value;
  var grid=document.getElementById('grid-lp');
  var cards=Array.from(grid.querySelectorAll('.album-card'));
  var vis=0;
  cards.forEach(function(c){
    var qOk=!q||c.dataset.search.includes(q);
    var origOk=origemFilter.size===0||origemFilter.has(c.dataset.origem||'');
    var isBrazil=c.dataset.country==='brazil';
    var nacOk=nacionalFilter==='all'||(nacionalFilter==='nacional'&&isBrazil)||(nacionalFilter==='internacional'&&!isBrazil);
    var decOk=decadeFilter.size===0||decadeFilter.has(c.dataset.decade);
    var compilOk=compilFilter==='all'||(compilFilter==='comp'&&c.dataset.compilation==='1')||(compilFilter==='nocomp'&&c.dataset.compilation!=='1');
    var djOk=djFilter==='all'||(djFilter==='yes'&&['Sim','Parcial'].indexOf(c.dataset.dj)!==-1);
    var paOk=paFilter==='all'||(paFilter==='yes'&&['Sim','Em breve'].indexOf(c.dataset.pa)!==-1);
    var dupOk=dupFilter==='all'||(dupFilter==='dup'&&+c.dataset.copies>1);
    var recOk=recebidoFilter==='all'||(recebidoFilter==='sim'&&c.dataset.recebido==='Sim')||(recebidoFilter==='nao'&&c.dataset.recebido!=='Sim');
    var bpmOk=albumBpmOk(c,activeBpmFilter);
    var ok=qOk&&origOk&&nacOk&&decOk&&compilOk&&djOk&&paOk&&dupOk&&recOk&&bpmOk;
    c.classList.toggle('hidden',!ok);if(ok)vis++;
  });
  document.getElementById('cnt-lp').textContent=vis;
  if(s){
    var vc=cards.filter(function(c){return!c.classList.contains('hidden')});
    vc.sort(function(a,b){
      if(s==='bpm-asc') return(+a.dataset.minBpm||999)-(+b.dataset.minBpm||999);
      if(s==='bpm-desc')return(+b.dataset.minBpm||0)-(+a.dataset.minBpm||0);
      if(s==='year-asc') return(+a.dataset.year||0)-(+b.dataset.year||0);
      if(s==='year-desc')return(+b.dataset.year||0)-(+a.dataset.year||0);
      if(s==='az')return(a.dataset.artist||'').localeCompare(b.dataset.artist||'');
      if(s==='added-asc') return(a.dataset.dateAdded||'').localeCompare(b.dataset.dateAdded||'');
      if(s==='added-desc')return(b.dataset.dateAdded||'').localeCompare(a.dataset.dateAdded||'');
      return 0;
    });
    vc.forEach(function(c){grid.appendChild(c)});
  }
}

// ── TRACK VIEW ────────────────────────────────────────────────────────────────
var activeBpmFilter=new Set();

function setBpmFilter(val,el){
  if(val==='all'){
    activeBpmFilter.clear();
    document.querySelectorAll('.bpm-chip').forEach(function(c){c.classList.remove('active');c.classList.remove('multi-active')});
    document.querySelectorAll('.bpm-chip[data-val="all"]').forEach(function(c){c.classList.add('active')});
  }else{
    document.querySelectorAll('.bpm-chip[data-val="all"]').forEach(function(c){c.classList.remove('active')});
    if(activeBpmFilter.has(val)){
      activeBpmFilter.delete(val);
      document.querySelectorAll('.bpm-chip[data-val="'+val+'"]').forEach(function(c){c.classList.remove('multi-active')});
    }else{
      activeBpmFilter.add(val);
      document.querySelectorAll('.bpm-chip[data-val="'+val+'"]').forEach(function(c){c.classList.add('multi-active')});
    }
    if(activeBpmFilter.size===0){document.querySelectorAll('.bpm-chip[data-val="all"]').forEach(function(c){c.classList.add('active')});}
  }
  filterTracks();
}
function bpmFilterOk(bpm,hasBpm,filterSet){
  if(filterSet.size===0)return true;
  return Array.from(filterSet).some(function(f){
    if(f==='with')return hasBpm;
    if(f==='nobpm')return !hasBpm;
    if(!hasBpm)return false;
    if(f==='sub70')return bpm<70;
    if(f==='140plus')return bpm>=140;
    var p=f.split('-');return bpm>=+p[0]&&bpm<+p[1];
  });
}
function filterTracks(){
  var q=document.getElementById('q-faixas').value.toLowerCase().trim();
  var s=document.getElementById('sort-faixas').value;
  var grid=document.getElementById('grid-faixas');
  var rows=Array.from(grid.querySelectorAll('.track-row'));
  var vis=0;
  grid.classList.toggle('inc-mode',incFilterActive);
  rows.forEach(function(r){
    var bpm=+r.dataset.bpm||0;
    var hasBpm=r.dataset.hasbpm==='1';
    var bpmOk=bpmFilterOk(bpm,hasBpm,activeBpmFilter);
    var qOk=!q||r.dataset.search.includes(q);
    var isBrazil=r.dataset.country==='brazil';
    var nacOk=nacionalFilter==='all'||(nacionalFilter==='nacional'&&isBrazil)||(nacionalFilter==='internacional'&&!isBrazil);
    var decOk=decadeFilter.size===0||decadeFilter.has(r.dataset.decade);
    var compilOk=compilFilter==='all'||(compilFilter==='comp'&&r.dataset.compilation==='1')||(compilFilter==='nocomp'&&r.dataset.compilation!=='1');
    var origOk=origemFilter.size===0||origemFilter.has(r.dataset.origem||'');
    var djOk=djFilter==='all'||(djFilter==='yes'&&['Sim','Parcial'].indexOf(r.dataset.dj)!==-1);
    var paOk=paFilter==='all'||(paFilter==='yes'&&['Sim','Em breve'].indexOf(r.dataset.pa)!==-1);
    var recOk=recebidoFilter==='all'||(recebidoFilter==='sim'&&r.dataset.recebido==='Sim')||(recebidoFilter==='nao'&&r.dataset.recebido!=='Sim');
    var inctype=r.dataset.inctype||'';
    var incOk=!incFilterActive||(inctype==='nosp'||inctype==='nobpm'||inctype==='both');
    var ok=qOk&&bpmOk&&nacOk&&decOk&&compilOk&&origOk&&djOk&&paOk&&recOk&&incOk;
    r.classList.toggle('hidden',!ok);if(ok)vis++;
  });
  document.getElementById('cnt-faixas').textContent=vis;
  if(s){
    var vr=rows.filter(function(r){return!r.classList.contains('hidden')});
    vr.sort(function(a,b){
      var artCmp=(a.dataset.artist||'').localeCompare(b.dataset.artist||'','pt');
      var bpmA=+a.dataset.bpm||9999, bpmB=+b.dataset.bpm||9999;
      if(s==='bpm-asc'){
        if(incFilterActive){
          // Sem Spotify(c/BPM)=0 → Sem BPM(c/Spotify)=1 → Sem ambos=2
          var ia=a.dataset.inctype==='nosp'?0:(a.dataset.inctype==='nobpm'?1:2);
          var ib=b.dataset.inctype==='nosp'?0:(b.dataset.inctype==='nobpm'?1:2);
          if(ia!==ib)return ia-ib;
        }
        return (bpmA-bpmB)||artCmp;
      }
      if(s==='bpm-desc'){var ba=+a.dataset.bpm||0,bb=+b.dataset.bpm||0;return (bb-ba)||artCmp;}
      if(s==='az') return artCmp||(bpmA-bpmB);
      if(s==='year-asc') return (+a.dataset.year||0)-(+b.dataset.year||0)||artCmp;
      if(s==='year-desc') return (+b.dataset.year||0)-(+a.dataset.year||0)||artCmp;
      return 0;
    });
    vr.forEach(function(r){grid.appendChild(r)});
  }
}

// ── INCOMPLETAS FILTER (filters #grid-faixas in-place) ───────────────────────
function toggleIncompletas(){
  incFilterActive=!incFilterActive;
  var link=document.getElementById('incomplete-link');
  if(link)link.classList.toggle('active',incFilterActive);
  filterTracks();
}

// ── DUPLICADOS FILTER (results bar quick-toggle) ──────────────────────────────
function toggleDupFilter(){
  var wasActive=dupFilter==='dup';
  dupFilter=wasActive?'all':'dup';
  var link=document.getElementById('dup-link');
  if(link)link.classList.toggle('active',!wasActive);
  document.querySelectorAll('.dup-chip').forEach(function(c){c.classList.remove('active')});
  document.querySelectorAll('.dup-chip[data-val="'+dupFilter+'"]').forEach(function(c){c.classList.add('active')});
  filterLP();
}

// ── DETAILS EXPAND ────────────────────────────────────────────────────────────
function toggleDetails(row){
  if(row.classList.contains('hidden'))return;
  var d=row.querySelector('.c-details');
  if(d)d.classList.toggle('open');
}

// ── EMBED ──────────────────────────────────────────────────────────────────────
function _doEmbed(el,iframe){
  var trackRow=el.closest('.track-row');
  var trackLp=el.closest('.track');
  var w=document.createElement('div');w.className='embed-below';
  w.appendChild(iframe);
  if(trackRow){
    // Faixas compact: embed goes inside .c-details (opens it)
    var details=trackRow.querySelector('.c-details');
    if(details){details.classList.add('open');details.appendChild(w);}
    else{trackRow.appendChild(w);}
  }else if(trackLp){
    // LP accordion: insert below the track row within tracks-list
    trackLp.parentNode.insertBefore(w,trackLp.nextSibling);
  }else{
    el.parentNode.replaceChild(iframe,el);return;
  }
}
function loadEmbed(el,tid){
  var iframe=document.createElement('iframe');
  iframe.src='https://open.spotify.com/embed/track/'+tid+'?utm_source=generator&autoplay=1';
  iframe.width='100%';iframe.height='80';iframe.frameBorder='0';
  iframe.allow='autoplay; clipboard-write; encrypted-media; fullscreen; picture-in-picture';
  iframe.className='sp-embed';
  _doEmbed(el,iframe);
}
function loadDeezerEmbed(el,did){
  var iframe=document.createElement('iframe');
  iframe.src='https://widget.deezer.com/widget/light/track/'+did+'?autoplay=true';
  iframe.width='100%';iframe.height='80';iframe.frameBorder='0';
  iframe.allow='autoplay; clipboard-write; encrypted-media; fullscreen; picture-in-picture';
  iframe.className='sp-embed';
  _doEmbed(el,iframe);
}

// ── INIT ──────────────────────────────────────────────────────────────────────
var t1;document.getElementById('q-lp').addEventListener('input',function(){clearTimeout(t1);t1=setTimeout(filterLP,200)});
var t2;document.getElementById('q-faixas').addEventListener('input',function(){clearTimeout(t2);t2=setTimeout(filterTracks,200)});
document.getElementById('cnt-lp').textContent=document.querySelectorAll('#grid-lp .album-card').length;
document.getElementById('cnt-faixas').textContent=document.querySelectorAll('#grid-faixas .track-row').length;
// Populate incomplete count on load (don't render yet)
(function(){
  var all=Array.from(document.querySelectorAll('#grid-faixas .track-row'));
  var n=all.filter(function(r){return r.dataset.hasbpm==='0'||r.dataset.hasspotify==='0';}).length;
  var lk=document.getElementById('incomplete-link');
  if(lk)lk.textContent=n+' incompletas';
})();

// ── DISCOGS EDIT MODE ─────────────────────────────────────────────────────────
var editMode=false;
function _dcfg(){try{return JSON.parse(localStorage.getItem('discogs_cfg')||'{}')}catch(e){return{}}}
function _dSave(o){localStorage.setItem('discogs_cfg',JSON.stringify(o))}

function toggleEditMode(){
  var cfg=_dcfg();
  if(!cfg.token||!cfg.username){openSetupModal();return;}
  editMode=!editMode;
  document.body.classList.toggle('edit-mode',editMode);
  document.querySelectorAll('.pencil-mode-btn').forEach(function(b){b.classList.toggle('active',editMode);});
  if(editMode) refreshDiscogsFieldOptions(); // atualiza opções em background
}

function refreshDiscogsFieldOptions(){
  var cfg=_dcfg();
  if(!cfg.token||!cfg.username)return;
  var h={'Authorization':'Discogs token='+cfg.token,'User-Agent':'ColecaoDoAmsa/1.0'};
  fetch('https://api.discogs.com/users/'+cfg.username+'/collection/fields',{headers:h})
    .then(function(r){return r.json();})
    .then(function(fd){
      var fo={};
      (fd.fields||[]).forEach(function(f){
        if(f.options&&f.options.length){
          fo[f.name]=f.options.map(function(o){return typeof o==='string'?o:(o.value||String(o));});
        }
      });
      var c2=_dcfg(); c2.fieldOptions=fo; _dSave(c2);
    }).catch(function(){});
}

function openSetupModal(){
  var cfg=_dcfg();
  var m=document.getElementById('setup-modal');
  if(cfg.username)document.getElementById('setup-username').value=cfg.username;
  if(cfg.token)   document.getElementById('setup-token').value=cfg.token;
  if(cfg.folder)  document.getElementById('setup-folder').value=cfg.folder;
  m.classList.add('open');
}
function closeSetupModal(){document.getElementById('setup-modal').classList.remove('open');}

async function connectDiscogs(){
  var username=document.getElementById('setup-username').value.trim();
  var token=document.getElementById('setup-token').value.trim();
  var folder=document.getElementById('setup-folder').value.trim()||'1';
  var st=document.getElementById('setup-status');
  if(!username||!token){st.textContent='Preencha usuário e token.';return;}
  st.textContent='Conectando…';
  var h={'Authorization':'Discogs token='+token,'User-Agent':'ColecaoDoAmsa/1.0'};
  try{
    var r=await fetch('https://api.discogs.com/users/'+username,{headers:h});
    if(!r.ok)throw new Error('Token inválido ('+r.status+')');
    var fr=await fetch('https://api.discogs.com/users/'+username+'/collection/fields',{headers:h});
    if(!fr.ok)throw new Error('Erro campos ('+fr.status+')');
    var fd=await fr.json();
    var ids={};var fieldOptions={};
    (fd.fields||[]).forEach(function(f){
      ids[f.name]=f.id;
      if(f.options&&f.options.length){
        fieldOptions[f.name]=f.options.map(function(o){return typeof o==='string'?o:(o.value||String(o));});
      }
    });
    _dSave({username:username,token:token,folder:folder,fieldIds:ids,fieldOptions:fieldOptions});
    st.innerHTML='<span style="color:#2a7a2a">✓ Conectado! Campos: '+Object.keys(ids).join(', ')+'</span>';
    setTimeout(function(){
      closeSetupModal();
      editMode=true;
      document.body.classList.add('edit-mode');
      document.querySelectorAll('.pencil-mode-btn').forEach(function(b){b.classList.add('active');});
    },1200);
  }catch(e){st.innerHTML='<span style="color:#c0392b">✗ '+e.message+'</span>';}
}

function openCardEdit(card){
  document.querySelectorAll('.card-edit-form.open').forEach(function(f){
    if(f.closest('.album-card')!==card)f.classList.remove('open');
  });
  var form=card.querySelector('.card-edit-form');
  if(!form)return;
  form.classList.toggle('open');
  form.querySelector('.cef-status').textContent='';
  // Populate Loja select from Discogs field options stored in localStorage
  var cfg=_dcfg();
  var lojaOpts=(cfg.fieldOptions||{})['Origem']||[];
  var lojaSel=form.querySelector('select[data-field="Origem"]');
  if(lojaSel&&lojaOpts.length){
    var curVal=lojaSel.dataset.curVal||lojaSel.value||'';
    lojaSel.innerHTML='<option value="">—</option>'+
      lojaOpts.map(function(o){
        return '<option value="'+o+'"'+(o===curVal?' selected':'')+'>'+o+'</option>';
      }).join('');
    lojaSel.value=curVal;
  }
}
function closeCardEdit(card){
  var form=card.querySelector('.card-edit-form');
  if(form){form.classList.remove('open');form.querySelector('.cef-status').textContent='';}
}

async function saveCardEdit(card){
  var cfg=_dcfg();
  if(!cfg.token||!cfg.username){alert('Configure o token primeiro.');return;}
  var inst=card.dataset.instanceId;
  var rid =card.dataset.releaseId;
  var fld =cfg.folder||'1';
  var form=card.querySelector('.card-edit-form');
  var st  =form.querySelector('.cef-status');
  if(!inst||!rid){st.textContent='Dados insuficientes (instance_id/release_id).';return;}
  st.textContent='Salvando…';
  var inputs=form.querySelectorAll('.cef-input');
  var errs=[];
  var btn=form.querySelector('.cef-save');
  btn.disabled=true;
  for(var i=0;i<inputs.length;i++){
    var inp=inputs[i];
    var fname=inp.dataset.field;
    var fid=(cfg.fieldIds||{})[fname];
    if(!fid){errs.push(fname+':sem ID');continue;}
    var val=inp.tagName==='TEXTAREA'?inp.value:inp.value;
    try{
      var url='https://api.discogs.com/users/'+cfg.username+
              '/collection/folders/'+fld+
              '/releases/'+rid+
              '/instances/'+inst+
              '/fields/'+fid;
      var res=await fetch(url,{
        method:'POST',
        headers:{'Authorization':'Discogs token='+cfg.token,
                 'User-Agent':'ColecaoDoAmsa/1.0',
                 'Content-Type':'application/json'},
        body:JSON.stringify({value:val})
      });
      if(!res.ok&&res.status!==204)throw new Error('HTTP '+res.status);
    }catch(e){errs.push(fname+': '+e.message);}
  }
  btn.disabled=false;
  if(errs.length){
    st.innerHTML='<span style="color:#c0392b">✗ '+errs.join(' · ')+'</span>';
  }else{
    st.innerHTML='<span style="color:#2a7a2a">✓ Salvo no Discogs!</span>';
    var vals2={};
    form.querySelectorAll('.cef-input').forEach(function(inp){vals2[inp.dataset.field]=inp.value.trim();});
    refreshCardDisplayedFields(card,vals2);
    storeLocalOverride(rid,vals2);
    setTimeout(function(){closeCardEdit(card);},900);
  }
}

function storeLocalOverride(rid,vals){
  var all=JSON.parse(localStorage.getItem('discogs_edits')||'{}');
  all[rid]=Object.assign(all[rid]||{},vals);
  localStorage.setItem('discogs_edits',JSON.stringify(all));
}

function applyLocalOverrides(){
  var all=JSON.parse(localStorage.getItem('discogs_edits')||'{}');
  Object.keys(all).forEach(function(rid){
    var card=document.querySelector('[data-release-id="'+rid+'"]');
    if(!card)return;
    var vals=all[rid];
    refreshCardDisplayedFields(card,vals);
    var form=card.querySelector('.card-edit-form');
    if(form){
      Object.keys(vals).forEach(function(field){
        var inp=form.querySelector('[data-field="'+field+'"]');
        if(!inp)return;
        inp.value=vals[field];
        if(inp.tagName==='SELECT')inp.dataset.curVal=vals[field];
      });
    }
  });
}
document.addEventListener('DOMContentLoaded',applyLocalOverrides);

function refreshCardDisplayedFields(card,vals){
  var row=card.querySelector('.fields-row');
  if(!row)return;
  var items=[];
  if(vals['Origem'])  items.push('<span class="field-item"><strong>Loja:</strong> '+vals['Origem']+'</span>');
  if(vals['DJ']&&vals['DJ']!=='Não') items.push('<span class="field-item"><strong>Discotecar:</strong> '+vals['DJ']+'</span>');
  if(vals['PA']&&vals['PA']!=='Não') items.push('<span class="field-item"><strong>Trocar:</strong> '+vals['PA']+'</span>');
  if(vals['$'])       items.push('<span class="field-item"><strong>$:</strong> '+vals['$']+'</span>');
  if(vals['Recebido?']) items.push('<span class="field-item"><strong>Recebido:</strong> '+vals['Recebido?']+'</span>');
  row.innerHTML=items.join('');
  var notesEl=card.querySelector('.field-notes');
  if(notesEl){notesEl.textContent=vals['Notas']||'';notesEl.style.display=vals['Notas']?'':'none';}
  // Sync data-* attributes so filters work immediately after save
  if(vals['DJ']!==undefined)    card.dataset.dj=vals['DJ'];
  if(vals['PA']!==undefined)    card.dataset.pa=vals['PA'];
  if(vals['Recebido?']!==undefined) card.dataset.recebido=vals['Recebido?'];
}
"""


# ── SVG logos reutilizados nos render functions ──────────────────────────────
_SVG_DISCOGS_SM = (
    '<svg viewBox="0 0 100 100" width="14" height="14" fill="currentColor" style="flex-shrink:0">'
    '<circle cx="50" cy="50" r="50"/>'
    '<circle cx="50" cy="50" r="42" fill="none" stroke="white" stroke-width="3.5"/>'
    '<circle cx="50" cy="50" r="31" fill="none" stroke="white" stroke-width="3.5"/>'
    '<circle cx="50" cy="50" r="21" fill="none" stroke="white" stroke-width="3.5"/>'
    '<circle cx="50" cy="50" r="13" fill="white"/>'
    '<circle cx="50" cy="50" r="4" fill="currentColor"/>'
    '<polygon points="72,3 85,17 28,97 15,83" fill="white"/>'
    '</svg>'
)
_SVG_SPOTIFY_SM = (
    '<svg viewBox="0 0 24 24" width="14" height="14" fill="currentColor" style="flex-shrink:0">'
    '<path d="M12 0C5.4 0 0 5.4 0 12s5.4 12 12 12 12-5.4 12-12S18.66 0 12 0z'
    'M17.521 17.34c-.24.359-.66.48-1.021.24-2.82-1.74-6.36-2.101-10.561-1.141'
    '-.418.122-.779-.179-.899-.539-.12-.421.18-.78.54-.9 4.56-1.021 8.52-.6'
    ' 11.64 1.32.42.18.479.659.301 1.02zm1.44-3.3c-.301.42-.841.6-1.262.3'
    '-3.239-1.98-8.159-2.58-11.939-1.38-.479.12-1.02-.12-1.14-.6-.12-.48'
    '.12-1.021.6-1.141C9.6 9.9 15 10.561 18.72 12.84c.361.181.54.78.241 1.2z'
    'M19.08 10.62C15.24 8.4 8.82 8.16 5.16 9.301c-.6.179-1.2-.181-1.38-.721'
    '-.18-.601.18-1.2.72-1.381 4.26-1.26 11.28-1.02 15.721 1.621.539.3.719'
    ' 1.02.419 1.56-.299.421-1.02.599-1.559.3z"/>'
    '</svg>'
)

# ── Icon-only variants for compact track rows (no text labels) ────────────────
_SVG_DISCOGS_IC = (
    '<svg viewBox="0 0 100 100" width="15" height="15" style="flex-shrink:0">'
    '<circle cx="50" cy="50" r="50" fill="currentColor"/>'
    '<circle cx="50" cy="50" r="42" fill="none" stroke="white" stroke-width="4"/>'
    '<circle cx="50" cy="50" r="31" fill="none" stroke="white" stroke-width="4"/>'
    '<circle cx="50" cy="50" r="21" fill="none" stroke="white" stroke-width="4"/>'
    '<circle cx="50" cy="50" r="13" fill="white"/>'
    '<circle cx="50" cy="50" r="4" fill="currentColor"/>'
    '<polygon points="72,3 85,17 28,97 15,83" fill="white"/>'
    '</svg>'
)
_SVG_SPOTIFY_IC = (
    '<svg viewBox="0 0 24 24" width="15" height="15" style="flex-shrink:0">'
    '<path fill="#1DB954" d="M12 0C5.4 0 0 5.4 0 12s5.4 12 12 12 12-5.4 12-12S18.66 0 12 0z'
    'M17.521 17.34c-.24.359-.66.48-1.021.24-2.82-1.74-6.36-2.101-10.561-1.141'
    '-.418.122-.779-.179-.899-.539-.12-.421.18-.78.54-.9 4.56-1.021 8.52-.6'
    ' 11.64 1.32.42.18.479.659.301 1.02zm1.44-3.3c-.301.42-.841.6-1.262.3'
    '-3.239-1.98-8.159-2.58-11.939-1.38-.479.12-1.02-.12-1.14-.6-.12-.48'
    '.12-1.021.6-1.141C9.6 9.9 15 10.561 18.72 12.84c.361.181.54.78.241 1.2z'
    'M19.08 10.62C15.24 8.4 8.82 8.16 5.16 9.301c-.6.179-1.2-.181-1.38-.721'
    '-.18-.601.18-1.2.72-1.381 4.26-1.26 11.28-1.02 15.721 1.621.539.3.719'
    ' 1.02.419 1.56-.299.421-1.02.599-1.559.3z"/>'
    '</svg>'
)
_SVG_PLAY_IC = (
    '<svg viewBox="0 0 24 24" width="15" height="15" fill="currentColor" style="flex-shrink:0">'
    '<polygon points="6,3 20,12 6,21"/>'
    '</svg>'
)


def pastel_gradient(hex_color: str) -> str:
    """CSS for compact track-row: true pastel mix (colour blended with white)."""
    if not hex_color or len(hex_color) != 7 or not hex_color.startswith("#"):
        return ""
    try:
        r, g, b = int(hex_color[1:3], 16), int(hex_color[3:5], 16), int(hex_color[5:7], 16)
    except ValueError:
        return ""
    def mix(c, a): return round(c * a + 255 * (1 - a))
    rL, gL, bL = mix(r, .62), mix(g, .62), mix(b, .62)   # left  62% colour
    rR, gR, bR = mix(r, .38), mix(g, .38), mix(b, .38)   # right 38% colour
    def lin(c):
        c /= 255; return c / 12.92 if c <= 0.04045 else ((c + .055) / 1.055) ** 2.4
    lum = 0.2126 * lin(rL) + 0.7152 * lin(gL) + 0.0722 * lin(bL)
    tc  = "#111" if lum > 0.30 else "#fff"
    left  = f"#{rL:02x}{gL:02x}{bL:02x}"
    right = f"#{rR:02x}{gR:02x}{bR:02x}"
    return f"background:linear-gradient(105deg,{left} 0%,{right} 100%);--tc:{tc}"


def render_track_lp(row):
    """Renderiza uma faixa dentro da view LP."""
    bpm_f      = safe_float(row.get("bpm"))
    uri        = _clean(row.get("spotify_uri"))
    status     = str(row.get("status") or "REJEITADO")
    deezer_id  = _clean(row.get("deezer_id"))
    release_id = _clean(row.get("release_id"))

    bpm_txt = f"{bpm_f:.0f}" if bpm_f else "—"
    dot_cls = {"ACEITO":"s-ok","REVISAR":"s-rev"}.get(status,"s-rej")

    # Discogs button
    discogs_btn = ""
    if release_id:
        d_url = f"https://www.discogs.com/release/{release_id}"
        discogs_btn = f'<a class="trk-btn" href="{d_url}" target="_blank">Discogs</a>'

    # Spotify button
    sp_btn = ""
    if uri and uri != "nan" and "spotify" in uri:
        sp_tid = uri.split(":")[-1]
        sp_btn = f'<a class="trk-btn" href="spotify:track:{sp_tid}">Spotify</a>'

    # Ouvir button (embed trigger)
    ouvir_btn = ""
    if uri and uri != "nan" and "spotify" in uri and status == "ACEITO":
        sp_tid = uri.split(":")[-1]
        ouvir_btn = f'<div class="trk-btn trk-ouvir-btn" onclick="loadEmbed(this,\'{sp_tid}\')">&#9654; Ouvir</div>'
    elif deezer_id:
        ouvir_btn = f'<div class="trk-btn trk-ouvir-btn" onclick="loadDeezerEmbed(this,\'{deezer_id}\')">&#9654; Ouvir</div>'

    btn_col = (f'<div class="trk-btn-col">{discogs_btn}{sp_btn}{ouvir_btn}</div>'
               if (discogs_btn or sp_btn or ouvir_btn) else "")

    return f'''<div class="track" data-bpm="{int(bpm_f) if bpm_f else 999}">
  <span class="trk-pos">{esc(row.get("position"))}</span>
  <span class="trk-status {dot_cls}"></span>
  <div class="trk-info">
    <div class="trk-artist">{esc(row.get("artist_clean"))}</div>
    <div class="trk-name">{esc(row.get("track_title"))}</div>
  </div>
  <div class="trk-badges">
    <span class="badge-bpm">{bpm_txt}</span>
  </div>
  {btn_col}
</div>'''


def render_album_lp(group, copy_count=1, fields=None, country="", color_pastel="", format_data=None):
    """Renderiza um card de álbum (LP view)."""
    fields    = fields or {}
    format_info = format_data or {}
    first     = group.iloc[0]
    cover     = esc(first.get("cover_url") or "")
    bpm_vals  = group["bpm"].apply(safe_float).dropna()
    min_bpm   = int(bpm_vals.min()) if len(bpm_vals) else 999
    bpm_list_str = ",".join(str(int(v)) for v in bpm_vals)
    if len(bpm_vals):
        bmin, bmax = int(bpm_vals.min()), int(bpm_vals.max())
        bpm_range = f"{bmin} BPM" if bmin == bmax else f"{bmin}–{bmax} BPM"
    else:
        bpm_range = ""
    n_ok      = (group["status"] == "ACEITO").sum()
    year_s    = str(int(first["year"])) if safe_float(first.get("year")) else ""
    styles_s  = esc(first.get("styles") or "")
    genres_s  = esc(first.get("genres") or "")
    release_id   = first.get("release_id", "")
    instance_id  = (fields.get("instance_id") or "").strip()
    country_s    = esc(country or "")
    fmt_label   = format_info.get("label", "") or ""
    fmt_size    = format_info.get("format_size", "") or ""
    is_compil_flag = "0"
    artist_lower   = (first.get("album_artist") or "").lower()
    if any(kw in artist_lower for kw in ["various", "v.a.", "variados", "aa.vv."]):
        is_compil_flag = "1"

    # Genre + Style unificados
    genre_style_s = _genre_style(genres_s, styles_s)

    # Decade key para filtro JS
    year_int  = int(first["year"]) if safe_float(first.get("year")) else 0
    decade_key = _decade(year_int)

    # Cor do card (Gradiente F: EE→55)
    card_style = ""
    if color_pastel and len(color_pastel) == 7:
        _, cvars = card_colors(color_pastel)
        if cvars:
            card_style = (f' style="background:linear-gradient(105deg,{color_pastel}EE 0%,'
                          f'{color_pastel}55 100%);{cvars}"')

    img_tag = (f'<img class="cover-img" src="{cover}" alt="" loading="lazy" '
               f'onerror="this.style.display=\'none\'">'
               if cover else '<div class="cover-ph">&#9836;</div>')

    meta_parts = []
    if year_s:       meta_parts.append(year_s)
    if country_s:    meta_parts.append(country_s)
    if fmt_size and fmt_size not in ("LP","Other"): meta_parts.append(esc(fmt_size))
    if genre_style_s: meta_parts.append(genre_style_s[:60])
    if bpm_range:    meta_parts.append(bpm_range)
    if fmt_label:    meta_parts.append(esc(fmt_label[:28]))
    tags = " · ".join(meta_parts)

    _months = ["jan","fev","mar","abr","mai","jun","jul","ago","set","out","nov","dez"]
    date_added_raw = (fields.get("date_added") or "").strip()
    date_added_display = ""
    if date_added_raw:
        try:
            _d = date_added_raw[:10].split("-")
            date_added_display = f"{int(_d[2])} {_months[int(_d[1])-1]} {_d[0]}"
        except Exception:
            pass

    copy_badge = f'<span class="copy-badge">{copy_count} c&#243;pias</span>' if copy_count > 1 else ""

    discogs_url = f"https://www.discogs.com/release/{release_id}"
    # Spotify link: primeiro track ACEITO do álbum
    _sp_ids = group[group["status"] == "ACEITO"]["track_id"].dropna()
    _sp_tid = str(_sp_ids.iloc[0]) if len(_sp_ids) > 0 else ""
    _sp_tid = "" if _sp_tid in ("nan", "None", "") else _sp_tid
    sp_album_link = f"spotify:track:{_sp_tid}" if _sp_tid else ""
    discogs_card_btn = (f'<a class="btn-card" href="{discogs_url}" target="_blank" '
                        f'onclick="event.stopPropagation()">'
                        f'{_SVG_DISCOGS_SM} Discogs</a>')
    spotify_card_btn = (
        f'<a class="btn-card" href="{sp_album_link}" '
        f'onclick="event.stopPropagation()">'
        f'{_SVG_SPOTIFY_SM} Spotify</a>'
    ) if sp_album_link else ""
    # edit_btn_html defined below (after instance_id is computed from edit_form_html section)
    # btn_group will be assembled after edit_btn_html is ready

    custom_html = ""
    field_items = []
    if fields.get("Origem"):   field_items.append(f'<span class="field-item"><strong>Loja:</strong> {esc(fields["Origem"])}</span>')
    if fields.get("DJ") and fields["DJ"] != "Não":   field_items.append(f'<span class="field-item"><strong>Discotecar:</strong> {esc(fields["DJ"])}</span>')
    if fields.get("PA") and fields["PA"] != "Não":   field_items.append(f'<span class="field-item"><strong>Trocar:</strong> {esc(fields["PA"])}</span>')
    if fields.get("$"):        field_items.append(f'<span class="field-item"><strong>$:</strong> {esc(fields["$"])}</span>')
    if fields.get("Recebido?"): field_items.append(f'<span class="field-item"><strong>Recebido:</strong> {esc(fields["Recebido?"])}</span>')
    mc = fields.get("Media Condition",""); sc = fields.get("Sleeve Condition","")
    if mc or sc:
        field_items.append(f'<span class="field-item"><strong>Cond:</strong> {esc(" / ".join(filter(None,[mc,sc])))}</span>')
    if field_items:
        custom_html += f'<div class="fields-row">{"".join(field_items)}</div>'
    if fields.get("Notas"):
        custom_html += f'<div class="field-notes">{esc(fields["Notas"])}</div>'

    track_titles = " ".join(str(r.get("track_title","")) for _, r in group.iterrows())
    search_str = html_module.escape(
        f'{first.get("album_artist","")} {first.get("album_title","")} '
        f'{styles_s} {genres_s} {year_s} {country} {track_titles} '
        f'{fmt_label} {fields.get("Origem","")} {fields.get("Notas","")}'.lower()
    )

    group_dedup = group.drop_duplicates(subset=["position"])
    tracks_html = "\n".join(render_track_lp(r) for _, r in group_dedup.iterrows())
    n_bpm = int(group_dedup["bpm"].apply(safe_float).notna().sum())
    n_preview = 0
    for _, _r in group_dedup.iterrows():
        _uri = str(_r.get("spotify_uri") or "")
        _did = str(_r.get("deezer_id") or "")
        _did = "" if _did in ("nan", "None", "none", "") else _did
        if (_uri and "spotify" in _uri and str(_r.get("status", "")) == "ACEITO") or _did:
            n_preview += 1
    _tinfo = [f"{n_ok} de {len(group_dedup)} faixas"]
    if n_bpm > 0:  _tinfo.append(f"{n_bpm} com BPM")
    if n_preview > 0: _tinfo.append(f"{n_preview} com pr&#233;via")
    alb_tracks_info = " &middot; ".join(_tinfo)
    origem_val = (fields.get("Origem") or "").strip().lower()
    country_key = country.strip().lower()

    # ── Inline edit form (only rendered when instance_id available) ─────────
    def _sel(fname, options):
        cur = (fields.get(fname) or "").strip()
        opts = f'<option value="">—</option>' + "".join(
            f'<option value="{o}"{" selected" if cur == o else ""}>{o}</option>'
            for o in options
        )
        return f'<select class="cef-input" data-field="{fname}">{opts}</select>'

    edit_form_html = ""
    if instance_id:
        cur_origem = esc((fields.get("Origem") or "").strip())
        loja_opt   = f'<option value="{cur_origem}" selected>{cur_origem}</option>' if cur_origem else ""
        edit_form_html = f'''<div class="card-edit-form">
  <div class="cef-grid">
    <span class="cef-label">Loja</span><select class="cef-input" data-field="Origem" data-cur-val="{cur_origem}"><option value="">—</option>{loja_opt}</select>
    <span class="cef-label">Pre&#231;o ($)</span><input class="cef-input" data-field="$" type="text" value="{esc((fields.get('$') or '').strip())}" placeholder="ex: 450">
    <span class="cef-label">Recebido?</span>{_sel("Recebido?",["Sim","Não"])}
    <span class="cef-label">Discotecar</span>{_sel("DJ",["Sim","Parcial","Não"])}
    <span class="cef-label">Trocar</span>{_sel("PA",["Sim","Em breve","Não"])}
    <span class="cef-label">Notas</span><textarea class="cef-input cef-textarea" data-field="Notas" rows="2" placeholder="Notas livres...">{esc((fields.get('Notas') or '').strip())}</textarea>
  </div>
  <div class="cef-actions">
    <span class="cef-status"></span>
    <button class="cef-btn cef-save" onclick="event.stopPropagation();saveCardEdit(this.closest('.album-card'))">Salvar</button>
    <button class="cef-btn cef-cancel" onclick="event.stopPropagation();closeCardEdit(this.closest('.album-card'))">Cancelar</button>
  </div>
</div>'''

    edit_btn_html = (
        f'<a class="btn-card cef-edit-btn" href="#" '
        f'onclick="event.preventDefault();event.stopPropagation();openCardEdit(this.closest(\'.album-card\'))">&#9998; Editar</a>'
        if instance_id else ""
    )
    btn_group = f'<div class="rg-col-center">{edit_btn_html}{discogs_card_btn}{spotify_card_btn}</div>'

    return f'''<article class="album-card"{card_style}
  data-search="{search_str}"
  data-artist="{esc(first.get('album_artist',''))}"
  data-year="{year_s}"
  data-min-bpm="{min_bpm}"
  data-origem="{esc(origem_val)}"
  data-country="{esc(country_key)}"
  data-decade="{decade_key}"
  data-compilation="{is_compil_flag}"
  data-format="{esc(fmt_size)}"
  data-dj="{esc((fields.get('DJ') or '').strip())}"
  data-pa="{esc((fields.get('PA') or '').strip())}"
  data-recebido="{esc((fields.get('Recebido?') or '').strip())}"
  data-bpm-list="{bpm_list_str}"
  data-copies="{copy_count}"
  data-date-added="{esc(date_added_raw)}"
  data-release-id="{release_id}"
  data-instance-id="{instance_id}">
  {f'<div class="cover-blur" style="background-image:url(\'{cover}\')"></div>' if cover else ''}
  <header class="album-header" onclick="toggleAlbum(this)">
    <div class="cover-wrap">{img_tag}</div>
    <div class="album-info">
      <div class="alb-artist">{esc(first.get("album_artist",""))}{copy_badge}</div>
      <div class="alb-title">{esc(first.get("album_title",""))}</div>
      {f'<div class="alb-meta">{tags}</div>' if tags else ''}
      {custom_html}
      <div class="alb-tracks-info">{alb_tracks_info}</div>
      {f'<div class="alb-date-added">adicionado {date_added_display}</div>' if date_added_display else ''}
    </div>
    {btn_group}
    <button class="toggle-btn" aria-label="expandir">&#8964;</button>
  </header>
  {edit_form_html}
  <div class="tracks-list collapsed">{tracks_html}</div>
</article>'''


def render_track_row(row, country="", color_pastel="", format_data=None, origem="", dj="", recebido="", pa="", notas=""):
    """Renderiza uma linha de faixa (Track view)."""
    format_info = format_data or {}
    bpm_f     = safe_float(row.get("bpm"))
    uri       = _clean(row.get("spotify_uri"))
    status    = str(row.get("status") or "REJEITADO")
    thumb     = esc(row.get("thumb_url") or row.get("cover_url") or "")
    deezer_id  = _clean(row.get("deezer_id"))
    release_id = row.get("release_id", "")
    year_s     = str(int(row["year"])) if safe_float(row.get("year")) else ""
    styles_s   = str(row.get("styles") or "")
    genres_s   = str(row.get("genres") or "")
    fmt_size         = format_info.get("format_size", "") or ""
    fmt_label        = format_info.get("label", "") or ""
    fmt_is_compil    = "0"
    artist_lower     = (row.get("album_artist") or row.get("artist_clean") or "").lower()
    if any(kw in artist_lower for kw in ["various", "v.a.", "variados", "aa.vv."]):
        fmt_is_compil = "1"
    country_key = (country or "").strip().lower()
    origem_val  = (origem or "").strip().lower()
    year_int    = int(row["year"]) if safe_float(row.get("year")) else 0
    decade_key  = _decade(year_int)
    genre_style_s = _genre_style(genres_s, styles_s, limit=4)

    bpm_txt = f"{bpm_f:.0f}" if bpm_f else "—"
    bpm_int = int(bpm_f) if bpm_f else 0
    has_bpm = "1" if bpm_f else "0"
    _tr_tid = _clean(row.get("track_id"))
    _sp_ok = bool(_tr_tid and status == "ACEITO")
    has_spotify = "1" if _sp_ok else "0"
    # Incomplete type: nosp=BPM yes/Spotify no, nobpm=BPM no/Spotify yes, both=neither
    if not bpm_f and not _sp_ok:
        inc_type = "both"
    elif not bpm_f:
        inc_type = "nobpm"
    elif not _sp_ok:
        inc_type = "nosp"
    else:
        inc_type = ""
    inc_badge = {"nosp":  '<span class="h-flag f-sp  inc-badge">sem Spotify</span>',
                 "nobpm": '<span class="h-flag f-bpm inc-badge">sem BPM</span>',
                 "both":  '<span class="h-flag f-both inc-badge">sem ambos</span>'}.get(inc_type, "")

    # ── inline extra info (badges + notes) ───────────────────────
    _pa_v  = (pa  or "").strip()
    _dj_v  = (dj  or "").strip()
    _rec_v = (recebido or "").strip()
    _not_v = (notas or "").strip()
    _extra_parts = []
    if _pa_v  and _pa_v  != "Não": _extra_parts.append('<span class="c-badge c-badge-trocar">Trocar</span>')
    if _dj_v  and _dj_v  != "Não": _extra_parts.append('<span class="c-badge c-badge-disc">DJ</span>')
    if _rec_v == "Não":             _extra_parts.append('<span class="c-badge c-badge-nrec">N&#227;o recebido</span>')
    if _not_v:                      _extra_parts.append(f'<span class="c-notes-inline">{esc(_not_v)}</span>')
    c_extra = f'<div class="c-extra">{"".join(_extra_parts)}</div>' if _extra_parts else ""

    # ── pastel gradient ──────────────────────────────────────────
    pg = pastel_gradient(color_pastel)
    row_style = f' style="{pg}"' if pg else ""

    # ── BPM chip (left of cover) ─────────────────────────────────
    bpm_el = (f'<div class="c-bpm">{bpm_txt}<small>bpm</small></div>'
              if bpm_f else '<div class="c-bpm-none">—</div>')

    # ── Cover ────────────────────────────────────────────────────
    if thumb:
        img_tag = (f'<img class="c-thumb" src="{thumb}" alt="" loading="lazy" '
                   f'onerror="this.style.display=\'none\'">')
    else:
        ph_bg = f"background:{color_pastel}99" if color_pastel else "background:var(--bg2)"
        img_tag = f'<div class="c-ph" style="{ph_bg}">&#9836;</div>'

    # ── Icon buttons ─────────────────────────────────────────────
    discogs_url = f"https://www.discogs.com/release/{release_id}"

    discogs_btn = (
        f'<a class="ic-btn" href="{discogs_url}" target="_blank" '
        f'onclick="event.stopPropagation()" title="Discogs">{_SVG_DISCOGS_IC}</a>'
    ) if release_id else ""

    spotify_btn = (
        f'<a class="ic-btn" href="spotify:track:{_tr_tid}" '
        f'onclick="event.stopPropagation()" title="Spotify">{_SVG_SPOTIFY_IC}</a>'
    ) if _tr_tid and status == "ACEITO" else ""

    if uri and "spotify" in uri and status == "ACEITO":
        tid = uri.split(":")[-1]
        play_btn = (f'<button class="ic-btn" title="Ouvir" '
                    f'onclick="event.stopPropagation();loadEmbed(this,\'{tid}\')">'
                    f'{_SVG_PLAY_IC}</button>')
    elif deezer_id:
        play_btn = (f'<button class="ic-btn" title="Ouvir Deezer" '
                    f'onclick="event.stopPropagation();loadDeezerEmbed(this,\'{deezer_id}\')">'
                    f'{_SVG_PLAY_IC}</button>')
    else:
        play_btn = f'<button class="ic-btn" disabled title="Sem áudio">{_SVG_PLAY_IC}</button>'

    # ── Expandable details ───────────────────────────────────────
    album_s = esc(row.get("album_title") or "")
    fmt_label_s = esc(fmt_label[:22]) if fmt_label else ""
    det_parts = [p for p in [year_s, esc(country), esc(genre_style_s), fmt_label_s] if p]
    details_html = ""
    if album_s:
        details_html += f'<div class="tr-album">{album_s}</div>'
    if det_parts:
        details_html += f'<div class="tr-meta">{" &middot; ".join(det_parts)}</div>'

    # ── Search / filter data ─────────────────────────────────────
    search_str = html_module.escape(
        f'{row.get("track_title","")} {row.get("artist_clean","")} '
        f'{row.get("album_title","")} {country} {genres_s} {styles_s} {year_s} {fmt_label}'.lower()
    )
    artist_str = html_module.escape(str(row.get("artist_clean") or ""))

    return f'''<div class="track-row"{row_style}
  data-bpm="{bpm_int}" data-hasbpm="{has_bpm}" data-hasspotify="{has_spotify}"
  data-inctype="{inc_type}" data-year="{year_int}"
  data-dj="{esc(dj.strip())}" data-pa="{esc(pa.strip())}" data-recebido="{esc(recebido.strip())}"
  data-search="{search_str}" data-artist="{artist_str}"
  data-country="{esc(country_key)}" data-decade="{decade_key}"
  data-compilation="{fmt_is_compil}" data-format="{esc(fmt_size)}"
  data-origem="{esc(origem_val)}"
  onclick="toggleDetails(this)">
  {bpm_el}
  {img_tag}
  <div class="c-info">
    <div class="c-title">{esc(row.get("track_title"))}</div>
    <div class="c-artist">{esc(row.get("artist_clean"))}</div>
    {c_extra}
    {inc_badge}
  </div>
  <div class="c-btns">{discogs_btn}{spotify_btn}{play_btn}</div>
  <div class="c-details">{details_html}</div>
</div>'''


def load_collection_fields():
    """
    Carrega backup_collection_fields.csv (gerado por fetch_discogs_fields.py).
    Retorna:
      - fields_map: {release_id: {campo: valor}} com os campos da primeira instância
      - copy_counts: {release_id: n_cópias}
    """
    path = WORK_DIR / "backup_collection_fields.csv"
    if not path.exists():
        return {}, {}

    import pandas as pd
    cf = pd.read_csv(path, dtype=str).fillna("")
    fields_map  = {}
    copy_counts = {}

    for rid, group in cf.groupby("release_id"):
        copy_counts[str(rid)] = len(group)
        # Agrega campos: usa a primeira linha, mas mescla Notas de múltiplas instâncias
        row = group.iloc[0].to_dict()
        notas_all = " | ".join(r for r in group.get("Notas", pd.Series([])).tolist() if r.strip())
        if notas_all:
            row["Notas"] = notas_all
        fields_map[str(rid)] = row

    return fields_map, copy_counts


def load_country_map():
    """Retorna {release_id: country} de backup_country.csv."""
    path = WORK_DIR / "backup_country.csv"
    if not path.exists():
        return {}
    import pandas as pd
    df = pd.read_csv(path, dtype=str).fillna("")
    return {str(r["release_id"]): r.get("country", "") for _, r in df.iterrows()}


def load_cover_colors():
    """Retorna {release_id: color_hex_original} de backup_colors.csv."""
    path = WORK_DIR / "backup_colors.csv"
    if not path.exists():
        return {}
    import pandas as pd
    df = pd.read_csv(path, dtype=str).fillna("")
    return {str(r["release_id"]): r.get("color_hex", "") for _, r in df.iterrows()}


def load_format_map():
    """Retorna {release_id: {label, is_compilation, format_size, format_raw}} de backup_format.csv."""
    path = WORK_DIR / "backup_format.csv"
    if not path.exists():
        return {}
    import pandas as pd
    df = pd.read_csv(path, dtype=str).fillna("")
    return {str(r["release_id"]): dict(r) for _, r in df.iterrows()}


def load_playlist_url():
    """Retorna URL da playlist Spotify salva, ou string vazia."""
    path = WORK_DIR / "backup_playlist.txt"
    if path.exists():
        return path.read_text(encoding="utf-8").strip()
    return ""


def generate_html(df):
    print("\nGerando HTML...")
    path = WORK_DIR / "MinhaColecao_DJ.html"

    # Carrega campos, country, cores, formato e playlist
    fields_map, copy_counts = load_collection_fields()
    country_map  = load_country_map()
    colors_map   = load_cover_colors()
    format_map   = load_format_map()
    playlist_url = load_playlist_url()

    n_unique  = df["release_id"].nunique()
    n_items   = sum(copy_counts.get(str(rid), 1) for rid in df["release_id"].unique()) if copy_counts else n_unique
    n_dupes   = n_items - n_unique   # total extra copies
    n_tracks  = len(df.drop_duplicates(subset=["release_id","position"]))  # sem duplicatas de cópia
    dup_link  = (f'<a class="incomplete-link" id="dup-link" onclick="toggleDupFilter()" style="cursor:pointer">{n_dupes} duplicados</a>'
                 if n_dupes > 0 else f'{n_dupes} duplicados')
    n_matched = (df.drop_duplicates(subset=["release_id","position"])["status"] == "ACEITO").sum()
    pct       = round(n_matched / n_tracks * 100) if n_tracks else 0
    bpm_vals  = df["bpm"].apply(safe_float).dropna()
    has_bpm   = len(bpm_vals) > 0

    # ── LP view ──────────────────────────────────────────────────────────────
    albums_html = "\n".join(
        render_album_lp(
            g,
            copy_count   = copy_counts.get(str(rid), 1),
            fields       = fields_map.get(str(rid), {}),
            country      = country_map.get(str(rid), ""),
            color_pastel = colors_map.get(str(rid), ""),
            format_data  = format_map.get(str(rid), {}),
        )
        for rid, g in df.sort_values(["album_artist","album_title"]).groupby("release_id", sort=False)
    )

    # ── Track view: sorted by BPM (deduplica cópias) ─────────────────────────
    df_tracks = df.drop_duplicates(subset=["release_id","position"]).copy()
    df_tracks["_bpm_sort"] = df_tracks["bpm"].apply(
        lambda v: safe_float(v) if safe_float(v) else 9999
    )
    df_tracks = df_tracks.sort_values("_bpm_sort")
    tracks_html = "\n".join(
        render_track_row(
            r,
            country      = country_map.get(str(r.get("release_id","")), ""),
            color_pastel = colors_map.get(str(r.get("release_id","")), ""),
            format_data  = format_map.get(str(r.get("release_id","")), {}),
            origem       = (fields_map.get(str(r.get("release_id","")), {}) or {}).get("Origem", ""),
            dj           = (fields_map.get(str(r.get("release_id","")), {}) or {}).get("DJ", ""),
            recebido     = (fields_map.get(str(r.get("release_id","")), {}) or {}).get("Recebido?", ""),
            pa           = (fields_map.get(str(r.get("release_id","")), {}) or {}).get("PA", ""),
            notas        = (fields_map.get(str(r.get("release_id","")), {}) or {}).get("Notas", ""),
        )
        for _, r in df_tracks.iterrows()
    )

    bpm_notice = ""
    if not has_bpm:
        bpm_notice = ('<div style="background:#F5F0E8;border:1px solid #D8CCC0;'
                      'padding:.6rem 1.2rem;border-radius:3px;font-size:.78rem;'
                      'color:#5C4030;margin-bottom:1rem;">'
                      'BPM n&#227;o dispon&#237;vel — execute dj_library_v2.py novamente para buscar via Deezer.</div>')

    # ── Chip inner HTML (sem wrapper div, para uso nos filter-groups) ─────────
    nac_inner = (
        '<button class="chip nac-chip active" data-val="all"'
        ' onclick="setNacionalFilter(\'all\',this)">Tudo</button>'
        '<button class="chip nac-chip" data-val="nacional"'
        ' onclick="setNacionalFilter(\'nacional\',this)">Nacional</button>'
        '<button class="chip nac-chip" data-val="internacional"'
        ' onclick="setNacionalFilter(\'internacional\',this)">Internacional</button>'
    )

    decade_inner = (
        '<button class="chip decade-chip active" data-val="all"'
        ' onclick="setDecadeFilter(this.dataset.val,this)">Tudo</button>'
        + ''.join(
            f'<button class="chip decade-chip" data-val="{dk}"'
            f' onclick="setDecadeFilter(this.dataset.val,this)">{lbl}</button>'
            for dk, lbl in [
                ("pre70","1970-"),("70s","1970-80"),("80s","1980-90"),
                ("90s","1990-00"),("2000s","2000-10"),("2010s","2010-20"),("2020s","2020+"),
            ]
        )
    )

    compil_inner = (
        '<button class="chip compil-chip active" data-val="all"'
        ' onclick="setCompilFilter(\'all\',this)">Tudo</button>'
        '<button class="chip compil-chip" data-val="comp"'
        ' onclick="setCompilFilter(\'comp\',this)">Colet&#226;neas</button>'
        '<button class="chip compil-chip" data-val="nocomp"'
        ' onclick="setCompilFilter(\'nocomp\',this)">Simples</button>'
    )

    # Origem values
    origem_vals = sorted(set(
        (v.get("Origem") or "").strip()
        for v in fields_map.values()
        if (v.get("Origem") or "").strip()
    ))
    origem_inner = ""
    if origem_vals:
        origem_inner = (
            '<button class="chip origem-chip active" data-val="all"'
            ' onclick="setOrigemFilter(this.dataset.val,this)">Tudo</button>'
            + ''.join(
                f'<button class="chip origem-chip" data-val="{esc(v.lower())}"'
                f' onclick="setOrigemFilter(this.dataset.val,this)">{esc(v)}</button>'
                for v in origem_vals
            )
            + '<button class="chip origem-chip" data-val=""'
            '  onclick="setOrigemFilter(this.dataset.val,this)">Sem origem</button>'
        )

    bpm_chip_inner = ''.join(
        f'<button class="chip bpm-chip{" active" if v=="all" else ""}" '
        f'data-val="{v}" onclick="setBpmFilter(\'{v}\',this)">{lbl}</button>'
        for v, lbl in [
            ("all","Tudo"),("with","Com BPM"),
            ("sub70","70-"),("70-80","70-80"),("80-90","80-90"),
            ("90-100","90-100"),("100-110","100-110"),("110-120","110-120"),
            ("120-130","120-130"),("130-140","130-140"),
            ("nobpm","Sem BPM"),
        ]
    )

    # DJ / PA / Duplicados filter chips
    has_dj_data = any((v.get("DJ") or "").strip() for v in fields_map.values())
    dj_inner = (
        '<button class="chip dj-chip active" data-val="all" onclick="setDjFilter(\'all\',this)">Tudo</button>'
        '<button class="chip dj-chip" data-val="yes" onclick="setDjFilter(\'yes\',this)">Para DJ</button>'
    ) if has_dj_data else ''

    has_recebido_data = any((v.get("Recebido?") or "").strip() for v in fields_map.values())
    recebido_inner = (
        '<button class="chip recebido-chip active" data-val="all" onclick="setRecebidoFilter(\'all\',this)">Tudo</button>'
        '<button class="chip recebido-chip" data-val="sim" onclick="setRecebidoFilter(\'sim\',this)">Sim</button>'
        '<button class="chip recebido-chip" data-val="nao" onclick="setRecebidoFilter(\'nao\',this)">N&#227;o</button>'
    ) if has_recebido_data else ''

    has_pa_data = any((v.get("PA") or "").strip() for v in fields_map.values())
    pa_inner = (
        '<button class="chip pa-chip active" data-val="all" onclick="setPaFilter(\'all\',this)">Tudo</button>'
        '<button class="chip pa-chip" data-val="yes" onclick="setPaFilter(\'yes\',this)">Para trocar</button>'
    ) if has_pa_data else ''

    n_dup_releases = sum(1 for rid in df["release_id"].unique() if copy_counts.get(str(rid), 1) > 1)
    dup_inner = (
        '<button class="chip dup-chip active" data-val="all" onclick="setDupFilter(\'all\',this)">Tudo</button>'
        f'<button class="chip dup-chip" data-val="dup" onclick="setDupFilter(\'dup\',this)">'
        f'Duplicados ({n_dup_releases})</button>'
    )

    # ── Filter panels (collapsible) ───────────────────────────────────────────
    origem_group = (
        f'<div class="filter-group"><span class="filter-group-label">Loja</span>{origem_inner}</div>'
        if origem_inner else ''
    )

    dj_group       = (f'<div class="filter-group"><span class="filter-group-label">Para discotecar</span>{dj_inner}</div>'   if dj_inner       else '')
    pa_group       = (f'<div class="filter-group"><span class="filter-group-label">Para trocar</span>{pa_inner}</div>'       if pa_inner       else '')
    recebido_group = (f'<div class="filter-group"><span class="filter-group-label">Recebido?</span>{recebido_inner}</div>'  if recebido_inner else '')
    dup_group =  f'<div class="filter-group"><span class="filter-group-label">C&#243;pias</span>{dup_inner}</div>'

    _shared_groups = (
        f'<div class="filter-group"><span class="filter-group-label">D&#233;cada</span>{decade_inner}</div>'
        f'<div class="filter-group"><span class="filter-group-label">Origem</span>{nac_inner}</div>'
        f'<div class="filter-group"><span class="filter-group-label">Tipo</span>{compil_inner}</div>'
        f'{origem_group}'
        f'{dj_group}'
        f'{pa_group}'
        f'{recebido_group}'
        f'<div class="filter-group"><span class="filter-group-label">BPM</span>{bpm_chip_inner}</div>'
    )

    fp_lp     = f'<div class="filter-panel" id="fp-lp">{_shared_groups}</div>'
    fp_faixas = f'<div class="filter-panel" id="fp-faixas">{_shared_groups}</div>'

    # SVG logos para botões sociais do header
    SVG_SPOTIFY = (
        '<svg viewBox="0 0 24 24" width="15" height="15" fill="currentColor" style="flex-shrink:0">'
        '<path d="M12 0C5.4 0 0 5.4 0 12s5.4 12 12 12 12-5.4 12-12S18.66 0 12 0z'
        'M17.521 17.34c-.24.359-.66.48-1.021.24-2.82-1.74-6.36-2.101-10.561-1.141'
        '-.418.122-.779-.179-.899-.539-.12-.421.18-.78.54-.9 4.56-1.021 8.52-.6'
        ' 11.64 1.32.42.18.479.659.301 1.02zm1.44-3.3c-.301.42-.841.6-1.262.3'
        '-3.239-1.98-8.159-2.58-11.939-1.38-.479.12-1.02-.12-1.14-.6-.12-.48'
        '.12-1.021.6-1.141C9.6 9.9 15 10.561 18.72 12.84c.361.181.54.78.241 1.2z'
        'M19.08 10.62C15.24 8.4 8.82 8.16 5.16 9.301c-.6.179-1.2-.181-1.38-.721'
        '-.18-.601.18-1.2.72-1.381 4.26-1.26 11.28-1.02 15.721 1.621.539.3.719'
        ' 1.02.419 1.56-.299.421-1.02.599-1.559.3z"/>'
        '</svg>'
    )
    SVG_DISCOGS = (
        '<svg viewBox="0 0 100 100" width="15" height="15" fill="currentColor" style="flex-shrink:0">'
        '<circle cx="50" cy="50" r="50"/>'
        '<circle cx="50" cy="50" r="43" fill="none" stroke="white" stroke-width="2.5"/>'
        '<circle cx="50" cy="50" r="36" fill="none" stroke="white" stroke-width="2.5"/>'
        '<circle cx="50" cy="50" r="29" fill="none" stroke="white" stroke-width="2.5"/>'
        '<circle cx="50" cy="50" r="22" fill="none" stroke="white" stroke-width="2.5"/>'
        '<circle cx="50" cy="50" r="14" fill="white"/>'
        '<circle cx="50" cy="50" r="4" fill="currentColor"/>'
        '<polygon points="72,3 85,17 28,97 15,83" fill="white"/>'
        '</svg>'
    )

    # Botão Spotify (liga à playlist se disponível, senão ao Spotify genérico)
    spotify_href = playlist_url or "https://open.spotify.com"
    spotify_social_btn = (
        f'<a class="header-social-btn" href="{spotify_href}" target="_blank">'
        f'{SVG_SPOTIFY} Spotify</a>'
    )

    _og_base = "https://amsa2diop.github.io/colecaovinil"
    _og_desc = f"Biblioteca de vinis &#183; {n_items} discos &#183; {n_tracks} faixas"

    html = f"""<!DOCTYPE html>
<html lang="pt-BR">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Cole&#231;&#227;o do Amsa</title>
<!-- Open Graph / WhatsApp preview -->
<meta property="og:type"        content="website">
<meta property="og:url"         content="{_og_base}/">
<meta property="og:title"       content="Cole&#231;&#227;o do Amsa">
<meta property="og:description" content="{_og_desc}">
<meta property="og:image"       content="{_og_base}/preview.jpg">
<meta property="og:image:width" content="900">
<meta property="og:image:height" content="1200">
<meta name="twitter:card"       content="summary_large_image">
<meta name="twitter:image"      content="{_og_base}/preview.jpg">
<style>{CSS}</style>
</head>
<body>

<!-- ═══ SETUP MODAL ═══ -->
<div class="setup-overlay" id="setup-modal">
  <div class="setup-box">
    <h3>&#9998; Configurar edi&#231;&#227;o Discogs</h3>
    <label>Usu&#225;rio Discogs</label>
    <input class="setup-input" id="setup-username" type="text" placeholder="seu_usuario" autocomplete="off">
    <label>Token pessoal &nbsp;<a href="https://www.discogs.com/settings/developers" target="_blank">(gerar aqui &#8599;)</a></label>
    <input class="setup-input" id="setup-token" type="password" placeholder="token..." autocomplete="off">
    <label>Pasta (folder_id &mdash; normalmente 1)</label>
    <input class="setup-input" id="setup-folder" type="number" value="1" min="1" style="width:80px">
    <div id="setup-status"></div>
    <div class="setup-actions">
      <button class="setup-btn-connect" onclick="connectDiscogs()">Conectar</button>
      <button class="setup-btn-cancel" onclick="closeSetupModal()">Cancelar</button>
    </div>
  </div>
</div>

<header class="site-header">
  <a class="logo-name logo-link" href="https://www.instagram.com/amsa2diop" target="_blank">Cole&#231;&#227;o do Amsa</a>
  <div class="header-sep"></div>
  <div class="site-stats">
    <span class="stat-item"><strong>{n_items}</strong> discos</span>
    <span class="stat-dot">&#8226;</span>
    <span class="stat-item"><strong>{n_tracks}</strong> faixas</span>
  </div>
  <div class="header-sep"></div>
  <a class="header-social-btn" href="https://www.discogs.com/pt_BR/user/amsa2diop/collection" target="_blank">{SVG_DISCOGS} Discogs</a>
  {spotify_social_btn}
  <div class="header-sep"></div>
  <div class="header-tabs">
    <button class="tab-btn active" data-v="lp" onclick="switchView('lp')">Discos</button>
    <button class="tab-btn" data-v="faixas" onclick="switchView('faixas')">Faixas</button>
  </div>
</header>

<!-- ═══════════════ LP VIEW ═══════════════ -->
<div id="view-lp" class="view active">
  <div class="controls">
    <div class="ctrl-row">
      <input class="ctrl-input" id="q-lp" type="search" placeholder="Buscar artista, &#225;lbum, ano, faixa, selo ou g&#234;nero...">
      <select class="ctrl-sel" id="sort-lp" onchange="filterLP()">
        <option value="added-desc">Adicionado (recente)</option>
        <option value="added-asc">Adicionado (antigo)</option>
        <option value="year-desc">Ano (recente)</option>
        <option value="year-asc">Ano (antigo)</option>
        <option value="" selected>Artista A&#8594;Z</option>
      </select>
      <button class="filter-toggle-btn" id="fp-btn-lp" onclick="toggleFilterPanel('lp')">&#9881; Filtros &#9662;</button>
      <button class="pencil-mode-btn" onclick="toggleEditMode()" title="Modo edi&#231;&#227;o">&#9998;</button>
      <button class="back-top-btn" onclick="window.scrollTo({{top:0,behavior:'smooth'}})" title="Voltar ao topo">&#8679;</button>
    </div>
    {fp_lp}
  </div>
  <main class="main">
    {bpm_notice}
    <div class="results-bar"><strong id="cnt-lp">{n_unique}</strong> &#250;nicos &nbsp;&#183;&nbsp; {dup_link}</div>
    <div class="albums-grid" id="grid-lp">{albums_html}</div>
  </main>
</div>

<!-- ═══════════════ TRACK VIEW ═══════════════ -->
<div id="view-faixas" class="view">
  <div class="controls">
    <div class="ctrl-row">
      <input class="ctrl-input" id="q-faixas" type="search" placeholder="Buscar artista, &#225;lbum, ano, faixa, selo ou g&#234;nero...">
      <select class="ctrl-sel" id="sort-faixas" onchange="filterTracks()">
        <option value="year-desc">Ano (recente)</option>
        <option value="year-asc">Ano (antigo)</option>
        <option value="az">Artista A&#8594;Z</option>
        <option value="bpm-asc" selected>BPM crescente</option>
        <option value="bpm-desc">BPM decrescente</option>
      </select>
      <button class="filter-toggle-btn" id="fp-btn-faixas" onclick="toggleFilterPanel('faixas')">&#9881; Filtros &#9662;</button>
      <button class="pencil-mode-btn" onclick="toggleEditMode()" title="Modo edi&#231;&#227;o">&#9998;</button>
      <button class="back-top-btn" onclick="window.scrollTo({{top:0,behavior:'smooth'}})" title="Voltar ao topo">&#8679;</button>
    </div>
    {fp_faixas}
  </div>
  <main class="main">
    <div class="results-bar">
      <strong id="cnt-faixas">{n_tracks}</strong> faixas
      &nbsp;&middot;&nbsp;
      <a class="incomplete-link" id="incomplete-link" onclick="toggleIncompletas()" style="cursor:pointer"></a>
    </div>
    <div class="track-rows" id="grid-faixas">{tracks_html}</div>
  </main>
</div>

<script>{JS}</script>
</body>
</html>"""

    path.write_text(html, encoding="utf-8")
    print(f"✓ {path.name} salvo ({path.stat().st_size // 1024} KB)")

    # Também escreve index.html (servido pelo GitHub Pages)
    index_path = WORK_DIR / "index.html"
    index_path.write_text(html, encoding="utf-8")
    print(f"✓ index.html sincronizado ({index_path.stat().st_size // 1024} KB)")

    return path


# ==============================================================================
# ENTRY POINT
# ==============================================================================
if __name__ == "__main__":
    sp, df_final = main()
    generate_xlsx(df_final)
    create_playlist(sp, df_final)
    generate_html(df_final)

    n_tracks  = len(df_final)
    n_matched = (df_final["status"] == "ACEITO").sum()
    n_albums  = df_final["release_id"].nunique()
    bpm_count = df_final["bpm"].apply(safe_float).notna().sum()

    print(f"""
CONCLUIDO!
  MinhaColecao_DJ.xlsx
  MinhaColecao_DJ.html
  Playlist Spotify: "Meu Discogs - por BPM"
  {n_tracks} faixas | {n_matched} no Spotify | {n_albums} albums | BPM: {bpm_count} faixas
""")
