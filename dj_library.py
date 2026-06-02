#!/usr/bin/env python3
"""
dj_library.py — Biblioteca DJ
Discogs + Spotify: coleta faixas, busca BPM, gera XLSX e HTML interativo.
"""

import sys, os, re, time, html as html_module, math, threading, webbrowser
import http.server, urllib.parse
from pathlib import Path

# Garante UTF-8 no Windows (evita UnicodeEncodeError com emojis/acentos)
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")

# ==============================================================================
# 0. DEPENDÊNCIAS
# ==============================================================================
import subprocess

def pip_install(*pkgs):
    subprocess.check_call([sys.executable, "-m", "pip", "install", "-q", *pkgs])

try:
    import discogs_client
except ImportError:
    print("Instalando discogs-client..."); pip_install("discogs-client")
    import discogs_client

try:
    import spotipy
    from spotipy.oauth2 import SpotifyOAuth
except ImportError:
    print("Instalando spotipy..."); pip_install("spotipy")
    import spotipy
    from spotipy.oauth2 import SpotifyOAuth

try:
    from thefuzz import fuzz
except ImportError:
    print("Instalando thefuzz..."); pip_install("thefuzz", "python-Levenshtein")
    from thefuzz import fuzz

try:
    from unidecode import unidecode
except ImportError:
    print("Instalando unidecode..."); pip_install("unidecode")
    from unidecode import unidecode

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Instalando openpyxl..."); pip_install("openpyxl")
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.utils import get_column_letter

import pandas as pd
import json, requests

# ==============================================================================
# 1. CONFIGURAÇÃO
# ==============================================================================
DISCOGS_USER   = "amsa2diop"
DISCOGS_TOKEN  = "FmUsfbDTXUFCniDqrTckrwBfJxvIljOGuMdygfVD"
SP_CLIENT_ID   = "1ab6d898c52d42a19b737f451ce31e2a"
SP_CLIENT_SEC  = "3c8b2f47049b44e2af6937ea835e1f2f"
SP_REDIRECT    = "http://127.0.0.1:1410/"
SP_SCOPE       = "playlist-modify-public playlist-modify-private"

# Limiares de aceitação (escala 0-100, igual ao fuzz ratio)
LIMIAR_ACEITO  = 72   # >= isto → ACEITO
LIMIAR_REVISAR = 55   # >= isto → REVISAR

WORK_DIR = Path(__file__).parent

# ==============================================================================
# 2. AUTENTICAÇÃO SPOTIFY (OAuth2 com servidor local na porta 1410)
# ==============================================================================
def spotify_auth_local_server():
    """Abre browser, captura o código OAuth via servidor local e retorna token."""
    sp_oauth = SpotifyOAuth(
        client_id     = SP_CLIENT_ID,
        client_secret = SP_CLIENT_SEC,
        redirect_uri  = SP_REDIRECT,
        scope         = SP_SCOPE,
        cache_path    = str(WORK_DIR / ".spotify_cache"),
        open_browser  = False,
    )

    # Tenta usar token em cache primeiro
    token_info = sp_oauth.get_cached_token()
    if token_info and not sp_oauth.is_token_expired(token_info):
        print("✅ Token Spotify em cache (válido).")
        return spotipy.Spotify(auth_manager=sp_oauth)

    auth_code_holder = [None]

    class OAuthHandler(http.server.BaseHTTPRequestHandler):
        def do_GET(self):
            parsed = urllib.parse.urlparse(self.path)
            params = urllib.parse.parse_qs(parsed.query)
            if "code" in params:
                auth_code_holder[0] = params["code"][0]
                self.send_response(200)
                self.send_header("Content-type", "text/html; charset=utf-8")
                self.end_headers()
                self.wfile.write(
                    b"<h2>&#10003; Autenticado! Pode fechar esta aba.</h2>"
                )
            else:
                self.send_response(400)
                self.end_headers()

        def log_message(self, *args):
            pass  # silencia log HTTP

    server = http.server.HTTPServer(("127.0.0.1", 1410), OAuthHandler)
    t = threading.Thread(target=server.handle_request, daemon=True)
    t.start()

    auth_url = sp_oauth.get_authorize_url()
    print(f"\n🔐 Abrindo navegador para autenticação Spotify...")
    webbrowser.open(auth_url)
    print("   (Se não abrir, acesse manualmente:)")
    print(f"   {auth_url}\n")

    t.join(timeout=120)
    server.server_close()

    if not auth_code_holder[0]:
        raise TimeoutError("Autenticação Spotify não completada em 120s.")

    sp_oauth.get_access_token(auth_code_holder[0])
    return spotipy.Spotify(auth_manager=sp_oauth)


# ==============================================================================
# 3. UTILITÁRIOS
# ==============================================================================
def normalize(text):
    """Normaliza texto para comparação: sem acentos, sem pontuação, lowercase."""
    if not text:
        return ""
    text = unidecode(str(text)).lower()
    text = re.sub(r"\(.*?\)|\[.*?\]", "", text)       # remove (...) e [...]
    text = re.sub(r"\s*[-–—]\s+.*$", "", text)         # remove "- subtítulo"
    text = re.sub(r"\bfeat\.?\b.*$|\bft\.?\b.*$", "", text)  # remove feat/ft
    text = re.sub(r"[^a-z0-9 ]", "", text)
    return text.strip()


def match_score(disc_track, disc_artist, sp_track, sp_artist, album_artist=""):
    """Score combinado: título (65%) + melhor artista (35%). Retorna 0-100."""
    dt = normalize(disc_track)
    da = normalize(disc_artist)
    aa = normalize(album_artist)
    st = normalize(sp_track)
    sa = normalize(sp_artist)

    score_track  = fuzz.token_sort_ratio(dt, st)
    score_artist = max(
        fuzz.token_sort_ratio(da, sa),
        fuzz.token_sort_ratio(aa, sa) if aa else 0,
    )
    score = score_track * 0.65 + score_artist * 0.35

    # Penalizações: versões indesejadas
    sp_lo = sp_track.lower()
    di_lo = disc_track.lower()
    if re.search(r"\blive\b|ao vivo|concert|en vivo", sp_lo) and \
       not re.search(r"live|ao vivo|concert", di_lo):
        score -= 20
    if re.search(r"\bremix\b|\bedit\b|rework|dub mix", sp_lo) and \
       not re.search(r"remix|edit|rework|dub", di_lo):
        score -= 15
    if re.search(r"remaster", sp_lo) and not re.search(r"remaster", di_lo):
        score -= 5

    return max(0, score)


def key_to_camelot(key, mode):
    """Converte key+mode do Spotify para notação Camelot Wheel."""
    major = ["8B","3B","10B","5B","12B","7B","2B","9B","4B","11B","6B","1B"]
    minor = ["5A","12A","7A","2A","9A","4A","11A","6A","1A","8A","3A","10A"]
    if key is None or math.isnan(key):
        return ""
    key = int(key)
    return major[key] if mode == 1 else minor[key]


def bpm_color(bpm):
    """Cor CSS baseada no BPM."""
    if not bpm or math.isnan(bpm):
        return "#555"
    bpm = float(bpm)
    if bpm < 80:   return "#8b5cf6"
    if bpm < 95:   return "#3b82f6"
    if bpm < 110:  return "#06b6d4"
    if bpm < 122:  return "#10b981"
    if bpm < 132:  return "#f59e0b"
    if bpm < 145:  return "#f97316"
    return "#ef4444"


# ==============================================================================
# 4. DISCOGS: COLETA DE ÁLBUNS E FAIXAS
# ==============================================================================
def get_discogs_collection(user_token, username):
    print("📀 Coletando coleção do Discogs...")
    d    = discogs_client.Client("DJLibrary/2.0", user_token=user_token)
    user = d.user(username)
    col  = user.collection_folders[0].releases

    albums = []
    tracks = []
    total  = col.count
    done   = 0

    for item in col:
        try:
            rel = item.release
            album_artist = rel.artists[0].name if rel.artists else "Unknown"
            album_title  = rel.title
            release_id   = rel.id
            year         = getattr(rel, "year", None)
            genres       = ", ".join(getattr(rel, "genres", []) or [])
            styles       = ", ".join(getattr(rel, "styles", []) or [])
            # Capa
            images    = getattr(rel, "images", []) or []
            cover_url = images[0].get("uri", "") if images else ""
            thumb_url = images[0].get("uri150", cover_url) if images else ""

            albums.append({
                "release_id":   release_id,
                "album_artist": album_artist,
                "album_title":  album_title,
                "year":         year,
                "genres":       genres,
                "styles":       styles,
                "cover_url":    cover_url,
                "thumb_url":    thumb_url,
            })

            for trk in rel.tracklist:
                if trk.type_ != "track" or not trk.position:
                    continue
                trk_artist = trk.artists[0].name if trk.artists else album_artist
                # Remove sufixo numérico de artistas Discogs "(2)", "(3)", etc.
                trk_artist_clean = re.sub(r"\s*\(\d+\)$", "", trk_artist)

                tracks.append({
                    "release_id":    release_id,
                    "album_artist":  album_artist,
                    "album_title":   album_title,
                    "year":          year,
                    "genres":        genres,
                    "styles":        styles,
                    "cover_url":     cover_url,
                    "thumb_url":     thumb_url,
                    "position":      trk.position,
                    "track_title":   trk.title,
                    "artist_raw":    trk_artist,
                    "artist_clean":  trk_artist_clean,
                })

            done += 1
            print(f"\r  {done}/{total} álbuns | {len(tracks)} faixas", end="", flush=True)
            time.sleep(0.5)  # Discogs rate limit

        except Exception as e:
            done += 1
            print(f"\n  ⚠️  Erro álbum #{done}: {e}")
            time.sleep(2)

    print(f"\n✅ {done} álbuns | {len(tracks)} faixas coletadas")
    return pd.DataFrame(albums), pd.DataFrame(tracks)


# ==============================================================================
# 5. SPOTIFY: MATCHING MELHORADO (multi-estratégia, 5 candidatos)
# ==============================================================================
def search_spotify_track(sp, artist, album_artist, title):
    """Busca no Spotify com 4 estratégias e avalia até 5 candidatos."""
    time.sleep(0.35)

    def run_search(query, limit=5):
        try:
            res = sp.search(q=query, type="track", limit=limit)
            items = res["tracks"]["items"]
            return items if items else None
        except Exception:
            time.sleep(1)
            return None

    # Estratégia 1: campos específicos
    items = run_search(f'artist:"{artist}" track:"{title}"')

    # Estratégia 2: texto livre
    if not items:
        time.sleep(0.2)
        items = run_search(f"{artist} {title}")

    # Estratégia 3: album_artist + title (coletâneas/V.A.)
    if not items and artist.lower() != album_artist.lower():
        time.sleep(0.2)
        items = run_search(f"{album_artist} {title}")

    # Estratégia 4: só o título
    if not items:
        time.sleep(0.2)
        items = run_search(f'track:"{title}"')

    empty = dict(spotify_uri=None, found_name=None, found_artist=None,
                 track_id=None, match_score=0.0, search_strategy="not_found")

    if not items:
        return empty

    best_score = -1
    best       = empty
    strategy   = "field" if items else "other"

    for item in items[:5]:
        sp_name   = item.get("name", "")
        sp_artist = item["artists"][0]["name"] if item.get("artists") else ""
        sc = match_score(title, artist, sp_name, sp_artist, album_artist)

        if sc > best_score:
            best_score = sc
            best = dict(
                spotify_uri     = item["uri"],
                found_name      = sp_name,
                found_artist    = sp_artist,
                track_id        = item["id"],
                match_score     = round(sc, 1),
                search_strategy = strategy,
            )

    return best


# ==============================================================================
# 6. BPM via Audio Features (lotes de 100)
# ==============================================================================
def get_audio_features(sp, track_ids):
    """Busca BPM e outros atributos em lotes de 100."""
    valid = [t for t in track_ids if t]
    if not valid:
        return {}

    print(f"🎵 Buscando BPM para {len(valid)} faixas (lotes de 100)...")
    results = {}
    chunks  = [valid[i:i+100] for i in range(0, len(valid), 100)]

    for i, chunk in enumerate(chunks):
        time.sleep(0.4)
        try:
            feats = sp.audio_features(chunk)
            for f in feats:
                if f:
                    results[f["id"]] = {
                        "bpm":          round(f["tempo"], 1),
                        "energy":       round(f["energy"], 2),
                        "danceability": round(f["danceability"], 2),
                        "valence":      round(f["valence"], 2),
                        "key":          f["key"],
                        "mode":         f["mode"],
                        "camelot":      key_to_camelot(f["key"], f["mode"]),
                    }
        except Exception as e:
            print(f"\n  ⚠️  Erro features lote {i+1}: {e}")
        print(f"\r  BPM: lote {i+1}/{len(chunks)} ✓", end="", flush=True)

    print(f"\n✅ BPM obtido para {len(results)} faixas")
    return results


# ==============================================================================
# 7. EXECUÇÃO PRINCIPAL
# ==============================================================================
def main():
    # ── Auth Spotify ──────────────────────────────────────────────────────────
    sp = spotify_auth_local_server()
    print("✅ Spotify autenticado.\n")

    # ── Discogs ───────────────────────────────────────────────────────────────
    backup_tracks = WORK_DIR / "backup_tracks.csv"
    if backup_tracks.exists():
        print(f"📂 Carregando backup Discogs: {backup_tracks}")
        df = pd.read_csv(backup_tracks)
        albums_df = df.drop_duplicates("release_id")
    else:
        albums_df, df = get_discogs_collection(DISCOGS_TOKEN, DISCOGS_USER)
        df.to_csv(backup_tracks, index=False)
        print(f"💾 Backup salvo: {backup_tracks}")

    print(f"\n📦 {len(albums_df)} álbuns | {len(df)} faixas\n")

    # ── Spotify matching ──────────────────────────────────────────────────────
    backup_matched = WORK_DIR / "backup_matched.csv"
    if backup_matched.exists():
        print(f"📂 Carregando backup matching: {backup_matched}")
        df_matched = pd.read_csv(backup_matched)
    else:
        print(f"🔍 Matching Spotify para {len(df)} faixas...")
        sp_cols = ["spotify_uri","found_name","found_artist","track_id","match_score","search_strategy"]
        for col in sp_cols:
            df[col] = None

        for i, row in df.iterrows():
            result = search_spotify_track(
                sp,
                row["artist_clean"],
                row["album_artist"],
                row["track_title"],
            )
            for k, v in result.items():
                df.at[i, k] = v

            if (i + 1) % 50 == 0:
                print(f"\r  {i+1}/{len(df)} faixas buscadas...", end="", flush=True)

        print(f"\r  {len(df)}/{len(df)} faixas buscadas ✓")

        df["status"] = df["match_score"].apply(
            lambda s: "ACEITO" if s >= LIMIAR_ACEITO
            else ("REVISAR" if s >= LIMIAR_REVISAR else "REJEITADO")
        )
        df_matched = df
        df_matched.to_csv(backup_matched, index=False)
        print(f"💾 Backup salvo: {backup_matched}")

    aceitos   = (df_matched["status"] == "ACEITO").sum()
    revisao   = (df_matched["status"] == "REVISAR").sum()
    rejeitado = (df_matched["status"] == "REJEITADO").sum()
    print(f"\n✅ Aceitos: {aceitos} | 👀 Revisar: {revisao} | ❌ Rejeitados: {rejeitado}\n")

    # ── BPM ───────────────────────────────────────────────────────────────────
    backup_final = WORK_DIR / "backup_final.csv"
    if backup_final.exists():
        print(f"📂 Carregando backup final: {backup_final}")
        df_final = pd.read_csv(backup_final)
    else:
        accepted_ids = df_matched[df_matched["status"] == "ACEITO"]["track_id"].dropna().tolist()
        bpm_map      = get_audio_features(sp, accepted_ids)

        for col in ["bpm","energy","danceability","valence","key","mode","camelot"]:
            df_matched[col] = None

        for i, row in df_matched.iterrows():
            tid = row.get("track_id")
            if tid and tid in bpm_map:
                for k, v in bpm_map[tid].items():
                    df_matched.at[i, k] = v

        df_final = df_matched
        df_final.to_csv(backup_final, index=False)
        print(f"💾 Backup final salvo: {backup_final}")

    return sp, df_final, albums_df


# ==============================================================================
# 8. XLSX
# ==============================================================================
def generate_xlsx(df):
    print("\n📊 Gerando XLSX...")
    path = WORK_DIR / "MinhaColecao_DJ.xlsx"
    wb   = openpyxl.Workbook()

    header_fill = PatternFill("solid", fgColor="1db954")
    header_font = Font(color="000000", bold=True)
    ok_fill     = PatternFill("solid", fgColor="d1fae5")
    warn_fill   = PatternFill("solid", fgColor="fef3c7")
    rej_fill    = PatternFill("solid", fgColor="fee2e2")

    cols = ["status","album_artist","artist_clean","track_title","album_title",
            "year","genres","styles","bpm","camelot","energy","danceability","valence",
            "match_score","found_artist","found_name","search_strategy","spotify_uri"]

    def write_sheet(ws, data, name):
        ws.title = name
        if data.empty:
            ws.append(["Nenhum dado."])
            return
        sub = data[[c for c in cols if c in data.columns]].copy()
        # Cabeçalho
        ws.append(list(sub.columns))
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        # Dados
        for _, row in sub.iterrows():
            ws.append([None if (isinstance(v, float) and math.isnan(v)) else v
                       for v in row])
        # Cor por status
        status_col = list(sub.columns).index("status") + 1 if "status" in sub.columns else None
        if status_col:
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                val = row[status_col - 1].value
                fill = ok_fill if val=="ACEITO" else (warn_fill if val=="REVISAR" else rej_fill)
                for cell in row:
                    cell.fill = fill
        # Largura automática
        for i, col in enumerate(ws.columns, 1):
            max_len = max((len(str(c.value or "")) for c in col), default=8)
            ws.column_dimensions[get_column_letter(i)].width = min(max_len + 2, 50)

    # Aba 1: todos
    ws1 = wb.active
    write_sheet(ws1, df.sort_values(["album_artist","album_title","position"],
                                    na_position="last"), "Todos (por album)")

    # Aba 2: aceitos por BPM
    ws2 = wb.create_sheet()
    df_ok = df[df["status"]=="ACEITO"].copy()
    df_ok["bpm"] = pd.to_numeric(df_ok["bpm"], errors="coerce")
    write_sheet(ws2, df_ok.sort_values("bpm", na_position="last"), "Aceitos (por BPM)")

    # Aba 3: para revisar
    ws3 = wb.create_sheet()
    write_sheet(ws3, df[df["status"]=="REVISAR"].sort_values("match_score", ascending=False),
                "Para Revisar")

    # Aba 4: rejeitados
    ws4 = wb.create_sheet()
    write_sheet(ws4, df[df["status"]=="REJEITADO"], "Nao Encontrados")

    wb.save(path)
    print(f"✅ {path.name} salvo!")
    return path


# ==============================================================================
# 9. PLAYLIST SPOTIFY (ordenada por BPM)
# ==============================================================================
def create_spotify_playlist(sp, df):
    print("\n🎧 Criando playlist Spotify (ordenada por BPM)...")
    df_pl = df[df["status"]=="ACEITO"].copy()
    df_pl["bpm"] = pd.to_numeric(df_pl["bpm"], errors="coerce")
    df_pl = df_pl.sort_values("bpm", na_position="last")
    uris  = df_pl["spotify_uri"].dropna().tolist()

    if not uris:
        print("⚠️  Nenhuma faixa aceita para playlist.")
        return

    from datetime import datetime
    now_str = datetime.now().strftime("%d/%m/%Y %H:%M")
    user_id = sp.current_user()["id"]

    pl = sp.user_playlist_create(
        user        = user_id,
        name        = "Meu Discogs — por BPM",
        public      = False,
        description = f"Coleção de vinis ordenada por BPM | {now_str}",
    )
    pid = pl["id"]

    chunks = [uris[i:i+100] for i in range(0, len(uris), 100)]
    for i, chunk in enumerate(chunks):
        sp.playlist_add_items(pid, chunk)
        time.sleep(0.3)
        print(f"\r  Lote {i+1}/{len(chunks)} enviado", end="", flush=True)

    print(f"\n🎉 Playlist criada com {len(uris)} faixas!")


# ==============================================================================
# 10. HTML: BIBLIOTECA DJ INTERATIVA
# ==============================================================================
def esc(v):
    return html_module.escape(str(v)) if v and str(v) != "nan" else ""


def track_html(row):
    bpm    = row.get("bpm")
    cam    = row.get("camelot", "") or ""
    energy = row.get("energy")
    dance  = row.get("danceability")
    uri    = row.get("spotify_uri", "")
    status = row.get("status", "REJEITADO")

    try:    bpm_f = float(bpm) if bpm and str(bpm) != "nan" else None
    except: bpm_f = None
    try:    ener_pct = int(float(energy) * 100) if energy and str(energy) != "nan" else 0
    except: ener_pct = 0
    try:    danc_pct = int(float(dance) * 100) if dance and str(dance) != "nan" else 0
    except: danc_pct = 0

    bpm_txt  = f"{bpm_f:.0f}" if bpm_f else "—"
    bpm_clr  = bpm_color(bpm_f) if bpm_f else "#555"
    cam_clr  = "#2563eb" if cam.endswith("A") else "#d97706"
    cam_span = f'<span class="cam-badge" style="background:{cam_clr}">{cam}</span>' if cam else ""

    dot_cls = {"ACEITO":"dot-ok","REVISAR":"dot-warn"}.get(status,"dot-rej")

    embed = ""
    if uri and str(uri) != "nan":
        tid = uri.split(":")[-1]
        embed = (f'<div class="embed-ph" onclick="loadEmbed(this,\'{tid}\')" title="Ouvir prévia">'
                 f'<span>&#9654;</span> Ouvir prévia</div>')
    else:
        embed = '<div class="no-embed">Não encontrado no Spotify</div>'

    return f'''<div class="track" data-bpm="{int(bpm_f) if bpm_f else 999}" data-status="{status}">
  <div class="track-main">
    <span class="status-dot {dot_cls}" title="{status}"></span>
    <div class="track-text">
      <div class="track-artist">{esc(row.get("artist_clean"))}</div>
      <div class="track-title">{esc(row.get("track_title"))}</div>
    </div>
    <div class="track-badges">
      <span class="bpm-badge" style="background:{bpm_clr}">{bpm_txt} BPM</span>{cam_span}
    </div>
    <div class="track-bars">
      <span class="bar-lbl">E</span><div class="bar-bg"><div class="bar-fill" style="width:{ener_pct}%;background:#f97316"></div></div>
      <span class="bar-lbl">D</span><div class="bar-bg"><div class="bar-fill" style="width:{danc_pct}%;background:#10b981"></div></div>
    </div>
  </div>
  {embed}
</div>'''


def album_html(group):
    first   = group.iloc[0]
    cover   = esc(first.get("cover_url", "") or "")
    bpm_vals = pd.to_numeric(group["bpm"], errors="coerce").dropna()
    bpm_range = f"{bpm_vals.min():.0f}–{bpm_vals.max():.0f} BPM" if len(bpm_vals) else "BPM n/d"
    min_bpm   = int(bpm_vals.min()) if len(bpm_vals) else 999
    n_ok      = (group["status"] == "ACEITO").sum()
    year_s    = str(int(first["year"])) if first.get("year") and str(first["year"]) != "nan" else ""
    genres_s  = esc(first.get("genres","") or "")
    styles_s  = esc(first.get("styles","") or "")

    img_tag = (f'<img class="cover-img" src="{cover}" alt="{esc(first.get("album_title",""))}" '
               f'loading="lazy" onerror="this.style.display=\'none\'">'
               if cover else '<div class="cover-ph">♪</div>')

    styles_tag = f'<span class="tag style-tag">{styles_s}</span>' if styles_s else ""
    tracks     = "\n".join(track_html(row) for _, row in group.iterrows())

    return f'''<article class="album-card" data-genres="{genres_s.lower()}" data-styles="{styles_s.lower()}" data-year="{year_s}" data-min-bpm="{min_bpm}">
  <header class="album-header" onclick="toggleAlbum(this)">
    <div class="cover-wrap">{img_tag}</div>
    <div class="album-info">
      <div class="alb-artist">{esc(first.get("album_artist",""))}</div>
      <div class="alb-title">{esc(first.get("album_title",""))}</div>
      <div class="alb-tags">
        <span class="tag year-tag">{year_s}</span>
        <span class="tag genre-tag">{genres_s}</span>{styles_tag}
      </div>
      <div class="alb-bpm">{bpm_range}</div>
      <div class="alb-ratio">{n_ok}/{len(group)} faixas no Spotify</div>
    </div>
    <button class="toggle-btn" aria-label="expandir">▼</button>
  </header>
  <div class="tracks-list collapsed">{tracks}</div>
</article>'''


CSS = """
:root{--bg:#09090f;--surf:#13131e;--surf2:#1c1c2c;--bdr:#2a2a3e;--acc:#1db954;--text:#e2e2e8;--muted:#6b6b90;--r:12px}
*{margin:0;padding:0;box-sizing:border-box}
body{font-family:"Segoe UI",system-ui,sans-serif;background:var(--bg);color:var(--text);min-height:100vh}
.site-header{background:linear-gradient(135deg,#0d1117 0%,#1a1a3e 60%,#0f3460 100%);padding:2rem 2.5rem 1.5rem;border-bottom:2px solid var(--acc)}
.site-header h1{font-size:2.2rem;color:var(--acc);font-weight:800;letter-spacing:-1px}
.site-header p{color:var(--muted);margin-top:.3rem;font-size:.9rem}
.stats-row{display:flex;gap:1.2rem;margin-top:1.2rem;flex-wrap:wrap}
.stat-card{background:rgba(255,255,255,.05);border:1px solid rgba(29,185,84,.25);border-radius:var(--r);padding:.6rem 1.2rem;text-align:center;min-width:85px}
.stat-val{font-size:1.5rem;font-weight:700;color:var(--acc);line-height:1}
.stat-lbl{font-size:.65rem;color:var(--muted);text-transform:uppercase;letter-spacing:.5px;margin-top:.2rem}
.controls{position:sticky;top:0;z-index:50;background:rgba(9,9,15,.95);backdrop-filter:blur(10px);border-bottom:1px solid var(--bdr);padding:.7rem 2rem;display:flex;gap:.7rem;flex-wrap:wrap;align-items:center}
.ctrl-input{flex:1;min-width:180px;background:var(--surf2);border:1px solid var(--bdr);color:var(--text);padding:.5rem .9rem;border-radius:8px;font-size:.9rem;outline:none}
.ctrl-input:focus{border-color:var(--acc)}
.ctrl-sel{background:var(--surf2);border:1px solid var(--bdr);color:var(--text);padding:.5rem .8rem;border-radius:8px;font-size:.85rem;outline:none;cursor:pointer}
.ctrl-btn{background:var(--surf2);border:1px solid var(--bdr);color:var(--text);padding:.5rem .9rem;border-radius:8px;font-size:.82rem;cursor:pointer;white-space:nowrap;transition:border-color .2s,color .2s}
.ctrl-btn:hover,.ctrl-btn.active{border-color:var(--acc);color:var(--acc);background:rgba(29,185,84,.08)}
.main{max-width:1300px;margin:0 auto;padding:1.5rem 2rem}
.results-bar{font-size:.83rem;color:var(--muted);margin-bottom:.8rem}
.albums-grid{display:flex;flex-direction:column;gap:1rem}
.album-card{background:var(--surf);border:1px solid var(--bdr);border-radius:var(--r);overflow:hidden;transition:border-color .2s}
.album-card:hover{border-color:rgba(29,185,84,.35)}
.album-card.hidden{display:none}
.album-header{display:flex;align-items:center;gap:1.2rem;padding:1rem 1.2rem;cursor:pointer;user-select:none;transition:background .15s}
.album-header:hover{background:rgba(255,255,255,.025)}
.cover-wrap{flex-shrink:0}
.cover-img{width:95px;height:95px;object-fit:cover;border-radius:8px;display:block}
.cover-ph{width:95px;height:95px;border-radius:8px;background:var(--surf2);display:flex;align-items:center;justify-content:center;font-size:2.2rem;color:var(--muted)}
.album-info{flex:1;min-width:0}
.alb-artist{font-weight:700;font-size:1rem}
.alb-title{color:var(--muted);font-size:.87rem;margin-top:.1rem;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}
.alb-tags{display:flex;flex-wrap:wrap;gap:.35rem;margin-top:.45rem}
.tag{padding:.18rem .55rem;border-radius:20px;font-size:.68rem;font-weight:600;text-transform:uppercase;letter-spacing:.3px}
.genre-tag{background:rgba(29,185,84,.12);color:#1db954;border:1px solid rgba(29,185,84,.25)}
.style-tag{background:rgba(59,130,246,.12);color:#60a5fa;border:1px solid rgba(59,130,246,.25)}
.year-tag{background:rgba(245,158,11,.12);color:#fbbf24;border:1px solid rgba(245,158,11,.25)}
.alb-bpm{font-size:.8rem;color:var(--muted);margin-top:.4rem}
.alb-ratio{font-size:.72rem;color:var(--muted);margin-top:.1rem}
.toggle-btn{background:none;border:none;color:var(--muted);font-size:1.1rem;cursor:pointer;padding:.3rem .5rem;transition:transform .3s,color .2s;flex-shrink:0}
.album-card.open .toggle-btn{transform:rotate(180deg);color:var(--acc)}
.tracks-list{border-top:1px solid var(--bdr)}
.tracks-list.collapsed{display:none}
.track{padding:.75rem 1.2rem;border-bottom:1px solid rgba(255,255,255,.04);transition:background .12s}
.track:last-child{border-bottom:none}
.track:hover{background:rgba(255,255,255,.02)}
.track-main{display:flex;align-items:center;gap:.75rem;flex-wrap:wrap}
.status-dot{width:7px;height:7px;border-radius:50%;flex-shrink:0;cursor:default}
.dot-ok{background:#1db954}.dot-warn{background:#f59e0b}.dot-rej{background:#ef4444}
.track-text{flex:1;min-width:120px}
.track-artist{font-size:.75rem;color:var(--muted)}
.track-title{font-size:.92rem;font-weight:600}
.track-badges{display:flex;gap:.35rem;align-items:center;flex-shrink:0}
.bpm-badge{padding:.22rem .55rem;border-radius:6px;font-size:.78rem;font-weight:700;color:#fff;letter-spacing:.3px}
.cam-badge{padding:.22rem .5rem;border-radius:6px;font-size:.76rem;font-weight:700;color:#fff}
.track-bars{display:flex;align-items:center;gap:.35rem;flex-shrink:0}
.bar-lbl{font-size:.62rem;color:var(--muted);width:12px;text-align:center}
.bar-bg{background:rgba(255,255,255,.08);border-radius:3px;height:5px;width:52px;overflow:hidden}
.bar-fill{height:100%;border-radius:3px}
.embed-ph{margin-top:.55rem;display:inline-flex;align-items:center;gap:.4rem;background:rgba(29,185,84,.08);border:1px solid rgba(29,185,84,.2);color:var(--acc);padding:.35rem .8rem;border-radius:6px;font-size:.8rem;cursor:pointer;transition:background .15s}
.embed-ph:hover{background:rgba(29,185,84,.15)}
.sp-embed{margin-top:.55rem;border-radius:8px;display:block}
.no-embed{margin-top:.55rem;font-size:.78rem;color:var(--muted);background:var(--surf2);padding:.4rem .7rem;border-radius:6px;display:inline-block}
@media(max-width:600px){.site-header{padding:1.2rem 1rem}.site-header h1{font-size:1.5rem}.controls{padding:.6rem 1rem}.main{padding:1rem}.cover-img,.cover-ph{width:70px;height:70px}}
"""

JS = """
var showMissing=true;
function toggleAlbum(hdr){var c=hdr.closest(".album-card"),l=c.querySelector(".tracks-list");c.classList.toggle("open");l.classList.toggle("collapsed")}
function expandAll(){document.querySelectorAll(".album-card:not(.hidden)").forEach(function(c){c.classList.add("open");c.querySelector(".tracks-list").classList.remove("collapsed")})}
function collapseAll(){document.querySelectorAll(".album-card").forEach(function(c){c.classList.remove("open");c.querySelector(".tracks-list").classList.add("collapsed")})}
function toggleMissing(){showMissing=!showMissing;document.getElementById("miss-btn").textContent=showMissing?"Ocultar não encontrados":"Mostrar não encontrados";applyFilters()}
function applyFilters(){
  var q=document.getElementById("q").value.toLowerCase().trim();
  var g=document.getElementById("gf").value.toLowerCase();
  var s=document.getElementById("sf").value;
  var grid=document.getElementById("grid");
  var cards=Array.from(grid.querySelectorAll(".album-card"));
  var vis=0;
  cards.forEach(function(c){
    var ok=((!q)||c.textContent.toLowerCase().includes(q))&&((!g)||(c.dataset.genres||"").includes(g))&&(showMissing||c.querySelector(".dot-ok"));
    c.classList.toggle("hidden",!ok);if(ok)vis++;
  });
  document.getElementById("cnt").textContent=vis;
  if(s){
    var vc=cards.filter(function(c){return!c.classList.contains("hidden")});
    vc.sort(function(a,b){
      if(s==="bpm-asc") return(+a.dataset.minBpm||999)-(+b.dataset.minBpm||999);
      if(s==="bpm-desc")return(+b.dataset.minBpm||0)-(+a.dataset.minBpm||0);
      if(s==="year-asc") return(+a.dataset.year||0)-(+b.dataset.year||0);
      if(s==="year-desc")return(+b.dataset.year||0)-(+a.dataset.year||0);
      if(s==="az")return a.querySelector(".alb-artist").textContent.localeCompare(b.querySelector(".alb-artist").textContent);
      return 0;
    });
    vc.forEach(function(c){grid.appendChild(c)});
  }
}
function loadEmbed(el,tid){var iframe=document.createElement("iframe");iframe.src="https://open.spotify.com/embed/track/"+tid+"?utm_source=generator";iframe.width="100%";iframe.height="80";iframe.frameBorder="0";iframe.allow="autoplay;clipboard-write;encrypted-media;fullscreen;picture-in-picture";iframe.className="sp-embed";el.parentNode.replaceChild(iframe,el)}
var timer;document.getElementById("q").addEventListener("input",function(){clearTimeout(timer);timer=setTimeout(applyFilters,220)});
document.getElementById("cnt").textContent=document.querySelectorAll(".album-card").length;
"""


def generate_html(df):
    print("🌐 Gerando HTML...")
    path = WORK_DIR / "MinhaColecao_DJ.html"

    n_albums  = df["release_id"].nunique()
    n_tracks  = len(df)
    n_matched = (df["status"] == "ACEITO").sum()
    bpm_vals  = pd.to_numeric(df["bpm"], errors="coerce").dropna()
    avg_bpm   = f"{bpm_vals.mean():.0f}" if len(bpm_vals) else "—"
    pct       = round(n_matched / n_tracks * 100) if n_tracks else 0

    genres_list = sorted(set(
        g.strip() for gs in df["genres"].dropna()
        for g in str(gs).split(",") if g.strip()
    ))
    genre_opts = "\n".join(
        f'<option value="{g.lower()}">{g}</option>' for g in genres_list
    )

    albums_html = "\n".join(
        album_html(group)
        for _, group in df.sort_values(["album_artist","album_title"]).groupby("release_id", sort=False)
    )

    html_str = f"""<!DOCTYPE html>
<html lang="pt-BR">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>DJ Library</title>
<style>{CSS}</style>
</head>
<body>
<header class="site-header">
  <h1>&#127911; Minha Biblioteca DJ</h1>
  <p>Cole&#231;&#227;o de vinis &bull; Discogs + Spotify</p>
  <div class="stats-row">
    <div class="stat-card"><div class="stat-val">{n_albums}</div><div class="stat-lbl">&#193;lbuns</div></div>
    <div class="stat-card"><div class="stat-val">{n_tracks}</div><div class="stat-lbl">Faixas</div></div>
    <div class="stat-card"><div class="stat-val">{n_matched}</div><div class="stat-lbl">No Spotify</div></div>
    <div class="stat-card"><div class="stat-val">{avg_bpm}</div><div class="stat-lbl">BPM m&#233;dio</div></div>
    <div class="stat-card"><div class="stat-val">{pct}%</div><div class="stat-lbl">Cobertura</div></div>
  </div>
</header>
<div class="controls">
  <input class="ctrl-input" id="q" type="search" placeholder="&#128269; Artista, t&#237;tulo, &#225;lbum...">
  <select class="ctrl-sel" id="gf" onchange="applyFilters()">
    <option value="">Todos os g&#234;neros</option>{genre_opts}
  </select>
  <select class="ctrl-sel" id="sf" onchange="applyFilters()">
    <option value="">Ordenar padr&#227;o</option>
    <option value="bpm-asc">BPM crescente</option>
    <option value="bpm-desc">BPM decrescente</option>
    <option value="year-asc">Ano (antigo)</option>
    <option value="year-desc">Ano (recente)</option>
    <option value="az">Artista A&#8594;Z</option>
  </select>
  <button class="ctrl-btn" onclick="expandAll()">Expandir tudo</button>
  <button class="ctrl-btn" onclick="collapseAll()">Recolher tudo</button>
  <button class="ctrl-btn" id="miss-btn" onclick="toggleMissing()">Ocultar n&#227;o encontrados</button>
</div>
<main class="main">
  <div class="results-bar">Mostrando <strong id="cnt">{n_albums}</strong> &#225;lbuns</div>
  <div class="albums-grid" id="grid">{albums_html}</div>
</main>
<script>{JS}</script>
</body>
</html>"""

    path.write_text(html_str, encoding="utf-8")
    print(f"✅ {path.name} salvo!")
    return path


# ==============================================================================
# ENTRY POINT
# ==============================================================================
if __name__ == "__main__":
    sp, df_final, albums_df = main()
    generate_xlsx(df_final)
    create_spotify_playlist(sp, df_final)
    generate_html(df_final)

    n_tracks  = len(df_final)
    n_matched = (df_final["status"] == "ACEITO").sum()
    n_albums  = df_final["release_id"].nunique()
    bpm_vals  = pd.to_numeric(df_final["bpm"], errors="coerce").dropna()
    avg_bpm   = f"{bpm_vals.mean():.0f}" if len(bpm_vals) else "—"

    print(f"""
🎉 CONCLUÍDO!
   📊 MinhaColecao_DJ.xlsx
   🌐 MinhaColecao_DJ.html
   🎧 Playlist Spotify: "Meu Discogs — por BPM"
   Total: {n_tracks} faixas | {n_matched} no Spotify | {n_albums} álbuns | BPM médio: {avg_bpm}
""")
